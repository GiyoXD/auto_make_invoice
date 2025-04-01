import openpyxl
import os
import math
import json # Added for JSON loading
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet # Optional: for type hinting
from openpyxl.styles import Alignment # Added for alignment
from typing import List, Dict, Any, Optional # Optional: for type hinting
import traceback # For printing detailed errors

def add_table_to_sheet(
    ws: Worksheet,
    start_row: int,
    header_rows: List[List[Any]],
    # label_data removed - labels are now part of data_rows
    data_rows: List[List[Any]],
    footer_config: Dict[str, Any]
    # columns_to_merge parameter removed
) -> int:
    """
    Inserts a table structure (header, data, footer with SUM formulas)
    into a worksheet at start_row, shifting existing rows down.
    Merges cells in the header appropriately (vertical and horizontal).
    Assumes labels like 'VENDOR#:' are included in the first column of data_rows.

    Args:
        ws: The openpyxl worksheet object to write to.
        start_row: The row number *before* which to insert the table.
        header_rows: List of lists for header rows (expects 2 rows for proper merging).
        data_rows: List of lists for the main data (including any labels in col 1).
        footer_config: Dict for footer settings. Example:
                       {'keywords': ["TOTAL:"], 'calculate_cols': [5, 6],
                        'pre_footer_rows': [[None,'Comment',None]], # Static rows *before* totals
                        'static_rows': [[None, 'Comment', None]]} # Static rows *after* totals

    Returns:
        The row number immediately after the inserted table.
        Returns start_row if an error occurs.
    """

    try:
        # --- Calculate total rows needed ---
        num_header_rows = len(header_rows)
        num_data_rows = len(data_rows)
        num_pre_footer_rows = len(footer_config.get('pre_footer_rows', []))
        num_main_footer_rows = 1 if footer_config.get('keywords') or footer_config.get('calculate_cols') else 0
        num_static_footer_rows = len(footer_config.get('static_rows', []))
        total_table_rows = num_header_rows + num_data_rows + num_pre_footer_rows + num_main_footer_rows + num_static_footer_rows

        if total_table_rows <= 0:
             print("Warning: No rows calculated for the table. Nothing inserted.")
             return start_row

        # --- Insert Blank Rows ---
        ws.insert_rows(start_row, amount=total_table_rows)

        # --- Write into the newly inserted blank rows ---
        current_row = start_row
        num_header_cols = 0
        if header_rows:
            num_header_cols = len(header_rows[0])

        # --- Write Header Rows ---
        header_start_row = current_row
        for row_idx, row_data in enumerate(header_rows):
            actual_row_num = header_start_row + row_idx
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                ws.cell(row=actual_row_num, column=col_idx, value=cell_value)
            current_row += 1 # Increment current_row after writing each header row

        # --- Apply Header Merging ---
        if num_header_rows == 2: # Specific logic for 2 header rows
             print(f"Applying merging to header rows {header_start_row} and {header_start_row + 1}...")
             header_row_1_data = header_rows[0]
             header_row_2_data = header_rows[1]
             r1 = header_start_row
             r2 = header_start_row + 1

             for col_idx_0based, val1 in enumerate(header_row_1_data):
                 col_idx_1based = col_idx_0based + 1
                 if col_idx_1based > num_header_cols: continue # Ensure within bounds

                 val2 = header_row_2_data[col_idx_0based] if col_idx_0based < len(header_row_2_data) else None

                 # Check for Horizontal Merge ("Quantity" example)
                 if val1 == "Quantity" and val2 == "PCS": # Condition based on example structure
                     try:
                         # Find index of "SF" which should be next
                         sf_idx_0based = header_row_2_data.index("SF", col_idx_0based + 1)
                         sf_col_idx_1based = sf_idx_0based + 1
                         print(f"Merging header horizontally Col {col_idx_1based} to {sf_col_idx_1based} in row {r1}")
                         ws.merge_cells(start_row=r1, start_column=col_idx_1based, end_row=r1, end_column=sf_col_idx_1based)
                         # Center align the merged "Quantity" cell
                         cell_to_align = ws.cell(row=r1, column=col_idx_1based)
                         cell_to_align.alignment = Alignment(horizontal='center', vertical='center')
                     except (ValueError, IndexError) as e:
                         print(f"Warning: Could not perform horizontal merge for 'Quantity'. Structure mismatch? Error: {e}")

                 # Check for Vertical Merge (Value in row 1, None/Empty in row 2)
                 elif val1 is not None and val1 != '' and (val2 is None or val2 == ''):
                     # Skip if this column was part of the horizontal merge handled above
                     # (e.g., don't vertically merge the cell *under* "Quantity")
                     is_under_quantity_merge = False
                     try:
                         q_idx = header_row_1_data.index("Quantity")
                         sf_idx = header_row_2_data.index("SF", q_idx + 1)
                         if q_idx <= col_idx_0based <= sf_idx:
                             is_under_quantity_merge = True
                     except ValueError: pass # Quantity or SF not found

                     if not is_under_quantity_merge:
                         print(f"Merging header vertically Col {col_idx_1based} from row {r1} to {r2}")
                         try:
                             ws.merge_cells(start_row=r1, start_column=col_idx_1based, end_row=r2, end_column=col_idx_1based)
                             # Center align the vertically merged cell
                             cell_to_align = ws.cell(row=r1, column=col_idx_1based)
                             cell_to_align.alignment = Alignment(vertical='center')
                         except Exception as merge_err:
                             print(f"Error merging header vertically Col {col_idx_1based}: {merge_err}")

        # --- Write Data Rows ---
        table_data_start_row = current_row # Data starts after headers
        for row_data in data_rows:
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 # Attempt numeric conversion
                 try:
                     if isinstance(cell_value, str) and (cell_value.replace('.', '', 1).isdigit() or (cell_value.startswith('-') and cell_value[1:].replace('.', '', 1).isdigit())):
                          cell_value_numeric = float(cell_value)
                          cell_value = int(cell_value_numeric) if cell_value_numeric.is_integer() else cell_value_numeric
                 except (ValueError, TypeError): pass
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1
        table_data_end_row = current_row - 1

        # --- Merge Cells Vertically in Data Rows --- REMOVED ---

        # --- Write Pre-Footer Static Rows ---
        pre_footer_rows = footer_config.get('pre_footer_rows', [])
        for static_row_data in pre_footer_rows:
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        # --- Write Main Footer Row (Totals) ---
        if num_main_footer_rows > 0:
            footer_row_content = [None] * num_header_cols
            footer_keywords = footer_config.get('keywords', [])
            if footer_keywords: footer_row_content[0] = footer_keywords[0]

            sum_col_indices = footer_config.get('calculate_cols', [])
            if table_data_start_row <= table_data_end_row:
                for col_index in sum_col_indices:
                    if 1 <= col_index <= num_header_cols:
                        col_letter = get_column_letter(col_index)
                        formula = f"=SUM({col_letter}{table_data_start_row}:{col_letter}{table_data_end_row})"
                        footer_row_content[col_index - 1] = formula
                    else: print(f"Warning: Footer SUM col index {col_index} out of range.")
            else: # No data rows
                for col_index in sum_col_indices:
                     if 1 <= col_index <= num_header_cols: footer_row_content[col_index - 1] = 0

            if header_rows: # Add item count
                try:
                    item_col_name = next((name for name in ['ITEM N°', 'Product Code'] if name in header_rows[0]), None)
                    if item_col_name:
                        item_no_col_index = header_rows[0].index(item_col_name)
                        footer_row_content[item_no_col_index] = f"{len(data_rows)} ITEMS"
                except (ValueError, IndexError): print("Warning: Could not determine item column for item count.")

            for col_idx, cell_value in enumerate(footer_row_content, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1

        # --- Write Post-Footer Static Rows ---
        static_footer_rows = footer_config.get('static_rows', [])
        for static_row_data in static_footer_rows:
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        return current_row

    except Exception as e:
        print(f"--- ERROR occurred while inserting table at row {start_row} ---")
        print(f"Error details: {e}")
        traceback.print_exc()
        return start_row


# --- Main Execution Logic ---
if __name__ == "__main__":

    json_input_filename = "test.json"
    output_filename = "tables_from_json_header_merged.xlsx" # Changed output name

    # --- 1. Load JSON Data ---
    if not os.path.exists(json_input_filename):
        print(f"Error: JSON input file '{json_input_filename}' not found."); exit()
    try:
        with open(json_input_filename, 'r', encoding='utf-8') as f: json_data = json.load(f)
        print(f"Successfully loaded data from '{json_input_filename}'")
    except Exception as e: print(f"Error loading JSON: {e}"); traceback.print_exc(); exit()

    # --- 2. Prepare Workbook ---
    if os.path.exists(output_filename):
        try: os.remove(output_filename); print(f"Removed previous output file: '{output_filename}'")
        except Exception as e: print(f"Warning: Could not remove file '{output_filename}': {e}")

    try:
        wb = openpyxl.Workbook()
        sheet_name = json_data.get('metadata', {}).get('worksheet_name', 'Inserted Tables Report')
        sheet_name = sheet_name[:31].replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace(':', '').replace('\\', '').replace('/', '')
        ws = wb.active; ws.title = sheet_name
        print(f"Created workbook with sheet: '{ws.title}'")
    except Exception as e: print(f"Error creating workbook: {e}"); traceback.print_exc(); exit()

    ws['A1'] = f"Report: {json_data.get('metadata', {}).get('workbook_filename', 'JSON Data')}"
    ws['A2'] = f"Sheet: {ws.title}"; ws['A3'] = ""

    # --- 3. Process and Insert Tables from JSON ---
    next_row_to_insert = 4

    if 'processed_tables_data' not in json_data or not isinstance(json_data['processed_tables_data'], dict):
         print("Error: JSON data missing 'processed_tables_data' dictionary."); exit()

    static_labels_in_data = ["VENDOR#:", "Des : LEATHER", "Case Qty :", "MADE IN CAMBODIA"]

    # columns_to_merge_in_table removed from here

    for table_id, table_data in json_data['processed_tables_data'].items():
        print(f"\n--- Processing Table ID: {table_id} ---")
        if not isinstance(table_data, dict) or not table_data:
            print(f"Warning: Skipping Table ID '{table_id}' - invalid/empty data."); continue

        # --- 3a. Define Header & Map JSON keys ---
        # This exact structure is needed for the header merging logic in add_table_to_sheet
        header_rows = [
            ["Mark & N°", "P.O N°", "ITEM N°", "Description", "Quantity", None, "N.W (kgs)", "G.W (kgs)", "CBM", "Unit", "Amount"],
            [None, None, None, None, "PCS", "SF", None, None, None, None, None]
        ]
        col_map = {
            "po": 2, "item": 3, "reference_code": 4, "pcs": 5, "sqft": 6,
            "net": 7, "gross": 8, "cbm": 9, "unit": 10, "amount": 11
        }
        expected_keys = list(col_map.keys())
        if not all(key in table_data for key in expected_keys):
             missing = [k for k in expected_keys if k not in table_data]
             print(f"Warning: Skipping Table ID '{table_id}'. Missing keys: {missing}"); continue
        num_cols = len(header_rows[0]) if header_rows else 0

        # --- 3b. Prepare data_rows (Integrate static labels) ---
        data_rows = []
        try:
            first_key = next(iter(table_data))
            num_rows = len(table_data[first_key])
            for i in range(num_rows):
                row = [None] * num_cols
                if i < len(static_labels_in_data): row[0] = static_labels_in_data[i]
                else: row[0] = None
                for key, col_idx in col_map.items():
                    if col_idx > num_cols: continue
                    if i < len(table_data[key]): row[col_idx - 1] = table_data[key][i]
                    else: row[col_idx - 1] = None
                data_rows.append(row)
        except StopIteration: print(f"Warning: Table '{table_id}' empty. Skipping."); continue
        except IndexError as ie: print(f"Error transforming data: {ie}"); traceback.print_exc(); continue
        except Exception as e: print(f"Error transforming data: {e}"); traceback.print_exc(); continue

        # --- 3c. Define Footer Configuration ---
        hs_code_row = [None] * num_cols
        hs_code_text = "HS.CODE: 4107.XX.XX"
        item_no_col_index = 2 # 0-based index for "ITEM N°" (Col 3)
        if item_no_col_index < num_cols: hs_code_row[item_no_col_index] = hs_code_text
        else: print("Warning: 'ITEM N°' column index out of bounds for HS Code.")

        footer_config = {
            'keywords': [f"TOTALS (Table {table_id}):"],
            'calculate_cols': [
                col_map['pcs'], col_map['sqft'], col_map['net'],
                col_map['gross'], col_map['cbm'], col_map['amount']
             ],
             'pre_footer_rows': [ hs_code_row ],
            'static_rows': [ [None] * num_cols ]
        }

        # --- 3d. Call add_table_to_sheet ---
        print(f"Inserting table for ID '{table_id}' before row {next_row_to_insert}...")
        try:
            # columns_to_merge argument removed from call
            next_row_after_insertion = add_table_to_sheet(
                ws,
                next_row_to_insert,
                header_rows,
                data_rows,
                footer_config
                # columns_to_merge=columns_to_merge_in_table # Removed
            )
            print(f"Table ID '{table_id}' finished. Next content starts at row {next_row_after_insertion}")
            next_row_to_insert = next_row_after_insertion + 1 # Add spacing
        except Exception as insert_error:
             print(f"--- FAILED to insert Table ID '{table_id}' ---"); print(f"Error: {insert_error}"); traceback.print_exc()
             next_row_to_insert += 1 # Increment minimally

    # --- 4. Save the final workbook ---
    try:
        wb.save(output_filename)
        print(f"\n--- Workbook saved successfully: '{output_filename}' ---")
        print(f"Full path: {os.path.abspath(output_filename)}")
    except Exception as e:
        print(f"\n--- ERROR saving workbook: {e} ---"); traceback.print_exc()
    finally:
        if wb:
             try: wb.close()
             except Exception: pass

    print("\n--- Script finished ---")
