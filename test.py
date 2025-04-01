# Combined script including Template Manager and Table Insertion functionality

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
# Import range utils for merge checking (from template_manager (1))
from openpyxl.utils.cell import range_boundaries

import logging
import os
import shutil
import math # Needed for ceiling function if calculating footer rows precisely

# Import necessary types from 'typing'
from typing import Dict, List, Any, Union, Optional, Tuple

# --- Attempt to import the handler (from template_manager (1)) ---
# This assumes excel_handler.py contains an ExcelHandler class
try:
    # If you have an ExcelHandler class in a separate file,
    # ensure it's accessible in your Python environment.
    # from excel_handler import ExcelHandler
    # For this combined script, we'll define a placeholder if not found
    # Replace the line below with your actual import if available
    ExcelHandler = None
    if ExcelHandler is None:
        logging.warning("ExcelHandler class not found or imported. Template loading functions requiring it will fail.")

except ImportError:
    logging.critical("CRITICAL ERROR: Could not import ExcelHandler. Make sure excel_handler.py is in the same directory or Python path.")
    ExcelHandler = None # Define as None to avoid NameError later if import fails

# --- Constants (from template_manager (1)) ---
DEFAULT_ANCHOR_TEXT = "Mark & Nº"
DEFAULT_CELLS_BELOW = 5
DEFAULT_CELLS_RIGHT = 10
COPY_SUFFIX = "_copy"
ANCHOR_KEY = 'anchor'

# --- Module Level Cache (from template_manager (1)) ---
_template_cache: Dict[str, Union[Optional[Dict[str, Any]], Optional[List[Optional[Dict[str, Any]]]]]] = {}


# --- Helper Functions (from template_manager (1)) ---

def _find_anchor_cell(worksheet: Worksheet, anchor_text: str) -> Optional[Cell]:
    """
    Finds the FIRST cell containing the anchor text during a row-by-row scan.
    If the cell found is part of a merged range, it returns the TOP-LEFT cell.
    Args:
        worksheet: The openpyxl worksheet object.
        anchor_text: The text to search for (case-sensitive).
    Returns:
        The top-left Cell object of the merge (or the cell itself) containing
        the text, otherwise None.
    """
    prefix = "[TemplateManager._find_anchor_cell]"
    # logging.debug(f"{prefix} Searching for FIRST cell containing anchor text '{anchor_text}' in sheet '{worksheet.title}'...") # REMOVED

    merge_map: Dict[str, str] = {}
    for mc_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mc_range.bounds
        top_left_coord = f"{get_column_letter(min_col)}{min_row}"
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                coord = f"{get_column_letter(c)}{r}"
                merge_map[coord] = top_left_coord

    for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            current_coord = cell.coordinate
            cell_value = cell.value

            if cell_value is None:
                is_top_left_of_merge = current_coord in merge_map and merge_map[current_coord] == current_coord
                if not is_top_left_of_merge:
                    continue

            try:
                if cell_value is not None and anchor_text in str(cell_value):
                    if current_coord in merge_map:
                        top_left_coord = merge_map[current_coord]
                        if current_coord == top_left_coord:
                            return cell
                        else:
                            tl_col_str, tl_row_str = coordinate_from_string(top_left_coord)
                            tl_col = column_index_from_string(tl_col_str)
                            tl_row = int(tl_row_str)
                            try:
                                if 1 <= tl_row <= worksheet.max_row and 1 <= tl_col <= worksheet.max_column:
                                    return worksheet.cell(row=tl_row, column=tl_col)
                                else:
                                     logging.warning(f"{prefix} Calculated top-left {top_left_coord} is out of sheet bounds. Returning originally found cell {current_coord}.") # KEPT WARNING
                                     return cell # Fallback
                            except Exception as cell_access_err: # Catch potential errors accessing cell
                                logging.warning(f"{prefix} Error accessing calculated top-left cell {top_left_coord}: {cell_access_err}. Returning originally found cell {current_coord}.") # KEPT WARNING
                                return cell
                    else:
                        return cell
            except TypeError:
                continue
            except Exception as e:
                logging.warning(f"{prefix} Unexpected error checking cell {current_coord} value '{cell_value}': {e}") # KEPT WARNING
                continue

    logging.warning(f"{prefix} Anchor text '{anchor_text}' not found in any cell (or merge top-left) in sheet '{worksheet.title}'.") # KEPT WARNING
    return None


def _extract_relative_cells(
    worksheet: Worksheet,
    anchor_cell: Cell, # Expected to be top-left if merged
    num_below: int,
    num_right: int
) -> Optional[Dict[str, Union[Optional[Dict[str, Any]], Optional[List[Optional[Dict[str, Any]]]]]]]:
    """
    Extracts properties from the anchor cell (inc. merge info) and cells relative.
    (Function implementation remains the same as before)
    """
    prefix = "[TemplateManager._extract_relative_cells]"
    if not anchor_cell:
        logging.error(f"{prefix} Invalid anchor_cell provided.") # KEPT ERROR
        return None

    relative_data: Dict[str, Union[Optional[Dict[str, Any]], Optional[List[Optional[Dict[str, Any]]]]]] = {}
    anchor_row = anchor_cell.row
    anchor_col = anchor_cell.column
    anchor_coord = anchor_cell.coordinate

    default_row_height = worksheet.sheet_format.defaultRowHeight if worksheet.sheet_format and worksheet.sheet_format.defaultRowHeight else 15.0
    default_col_width = worksheet.sheet_format.defaultColWidth if worksheet.sheet_format and worksheet.sheet_format.defaultColWidth else 8.43

    def _get_cell_props(cell: Cell, target_r: int, target_c: int, is_anchor: bool = False) -> Dict[str, Any]:
        """Extracts value, dimensions, border, alignment, and merge info (if anchor)."""
        props: Dict[str, Any] = {}
        try:
            props['value'] = cell.value
            rd = worksheet.row_dimensions.get(target_r)
            props['height'] = rd.height if rd and rd.height is not None else default_row_height
            col_letter = get_column_letter(target_c)
            cd = worksheet.column_dimensions.get(col_letter)
            props['width'] = cd.width if cd and cd.width is not None else default_col_width
            border_styles: Dict[str, Optional[str]] = {'left': None, 'right': None, 'top': None, 'bottom': None}
            if cell.has_style and cell.border:
                border_styles['left'] = cell.border.left.style if cell.border.left else None
                border_styles['right'] = cell.border.right.style if cell.border.right else None
                border_styles['top'] = cell.border.top.style if cell.border.top else None
                border_styles['bottom'] = cell.border.bottom.style if cell.border.bottom else None
            props['border'] = border_styles
            alignment_styles: Dict[str, Any] = {'horizontal': None, 'vertical': None, 'wrap_text': None}
            if cell.has_style and cell.alignment:
                alignment_styles['horizontal'] = cell.alignment.horizontal
                alignment_styles['vertical'] = cell.alignment.vertical
                alignment_styles['wrap_text'] = cell.alignment.wrap_text
            props['alignment'] = alignment_styles
            anchor_merge_rows = 1
            if is_anchor:
                merge_info = None
                for mc_range in worksheet.merged_cells.ranges:
                    if mc_range.min_row == target_r and mc_range.min_col == target_c:
                        rows_merged = mc_range.max_row - mc_range.min_row + 1
                        cols_merged = mc_range.max_col - mc_range.min_col + 1
                        if rows_merged > 1 or cols_merged > 1:
                            merge_info = {'rows': rows_merged, 'cols': cols_merged}
                            anchor_merge_rows = rows_merged
                        break # Found the merge range starting at this anchor cell
                if merge_info:
                    props['merge_info'] = merge_info
                props['_anchor_merge_rows'] = anchor_merge_rows # Store merge height even if 1x1
            return props
        except Exception as e:
            logging.error(f"{prefix} Error reading properties for cell at R={target_r}, C={target_c}: {e}", exc_info=False) # KEPT ERROR
            return {'error': f"Error reading cell properties: {e}"}

    anchor_props = _get_cell_props(anchor_cell, anchor_row, anchor_col, is_anchor=True)
    relative_data[ANCHOR_KEY] = anchor_props
    anchor_merge_height = 1
    if isinstance(anchor_props, dict) and 'error' not in anchor_props:
        anchor_merge_height = anchor_props.get('_anchor_merge_rows', 1)

    for i in range(1, num_below + 1):
        target_row = anchor_row + i
        target_col = anchor_col
        key_name = f'bottom_{i}'
        if target_row > worksheet.max_row or target_col > worksheet.max_column:
            logging.warning(f"{prefix} Target {key_name} (R={target_row}, C={target_col}) out of sheet bounds ({worksheet.max_row}x{worksheet.max_column}). Storing None.") # KEPT WARNING
            relative_data[key_name] = None
            continue
        try:
            cell_to_extract = worksheet.cell(row=target_row, column=target_col)
            relative_data[key_name] = _get_cell_props(cell_to_extract, target_row, target_col)
        except IndexError:
            logging.warning(f"{prefix} Could not access cell for {key_name} at (R={target_row}, C={target_col}) despite bounds check. Storing error.") # KEPT WARNING
            relative_data[key_name] = {'error': f"IndexError accessing cell at R={target_row}, C={target_col}"}
        except Exception as exc:
            logging.error(f"{prefix} Error getting props for {key_name} at (R={target_row}, C={target_col}): {exc}") # KEPT ERROR
            relative_data[key_name] = {'error': f"Error getting props: {exc}"}

    for i in range(1, num_right + 1):
        target_base_col = anchor_col + i
        key_name = f'right_{i}'
        cells_in_column: List[Optional[Dict[str, Any]]] = []
        for j in range(anchor_merge_height):
            target_row = anchor_row + j
            target_col = target_base_col
            if target_row > worksheet.max_row or target_col > worksheet.max_column:
                logging.warning(f"{prefix}     Cell at (R={target_row}, C={target_col}) for '{key_name}' (row offset {j}) out of sheet bounds ({worksheet.max_row}x{worksheet.max_column}). Storing None in list.") # KEPT WARNING
                cells_in_column.append(None)
                continue
            try:
                cell_to_extract = worksheet.cell(row=target_row, column=target_col)
                props = _get_cell_props(cell_to_extract, target_row, target_col)
                cells_in_column.append(props)
            except IndexError:
                logging.warning(f"{prefix}     IndexError accessing cell for '{key_name}' at (R={target_row}, C={target_col}). Storing error in list.") # KEPT WARNING
                cells_in_column.append({'error': f"IndexError accessing cell at R={target_row}, C={target_col}"})
            except Exception as cell_exc:
                logging.error(f"{prefix}     Unexpected error processing cell for '{key_name}' at (R={target_row}, C={target_col}): {cell_exc}. Storing error in list.") # KEPT ERROR
                cells_in_column.append({'error': f"Unexpected error processing cell: {cell_exc}"})
        relative_data[key_name] = cells_in_column

    return relative_data


# --- Template Manager Public Interface Functions (from template_manager (1)) ---

def load_template(
    workbook_path: str,
    sheet_identifier: Union[str, int],
    template_name: str,
    anchor_text: str = DEFAULT_ANCHOR_TEXT,
    num_below: int = DEFAULT_CELLS_BELOW,
    num_right: int = DEFAULT_CELLS_RIGHT,
    force_reload: bool = False,
    create_copy: bool = False
) -> bool:
    """
    Loads template using ExcelHandler, finds anchor (merge-aware), extracts properties.
    Requires ExcelHandler class to be available.
    Args:
        (Same args as before)
    Returns:
        True if template loaded/found, False otherwise.
    """
    prefix = f"[TemplateManager.load_template(name='{template_name}')]"

    # Check if ExcelHandler was imported successfully
    if ExcelHandler is None:
         logging.critical(f"{prefix} ExcelHandler class not available. Cannot load template.")
         return False

    if not force_reload and template_name in _template_cache:
        return True

    # --- File Check and Copy Logic (Remains Here) ---
    if not os.path.exists(workbook_path):
        logging.error(f"{prefix} Original workbook file not found: '{workbook_path}'.")
        return False

    path_to_load = workbook_path
    copied_path_created = None
    if create_copy:
        try:
            base, ext = os.path.splitext(workbook_path)
            path_to_load = f"{base}{COPY_SUFFIX}{ext}"
            if os.path.exists(path_to_load):
                try:
                    os.remove(path_to_load)
                except Exception as remove_err:
                    logging.warning(f"{prefix} Could not remove existing copy '{path_to_load}': {remove_err}.")
            shutil.copy2(workbook_path, path_to_load) # copy2 preserves metadata
            copied_path_created = path_to_load
        except Exception as copy_err:
            logging.error(f"{prefix} Failed to create copy of '{workbook_path}': {copy_err}", exc_info=True)
            return False
    # --- End File Check and Copy Logic ---

    excel_handler = None # Initialize handler variable
    try:
        # Instantiate the handler (will raise FileNotFoundError if path_to_load doesn't exist)
        logging.info(f"{prefix} Initializing ExcelHandler for '{path_to_load}'")
        excel_handler = ExcelHandler(path_to_load) # Assumes ExcelHandler exists

        # Load the specific sheet using the handler
        logging.info(f"{prefix} Asking ExcelHandler to load sheet: '{sheet_identifier}'")
        worksheet = excel_handler.load_sheet(sheet_name=sheet_identifier, data_only=True) # data_only=True for values

        # Check if sheet loading failed within the handler
        if not worksheet:
            logging.error(f"{prefix} ExcelHandler failed to load worksheet '{sheet_identifier}' from '{path_to_load}'.")
            return False
        else:
             logging.info(f"{prefix} ExcelHandler successfully loaded sheet: '{worksheet.title}'")


        # --- Find Anchor and Extract Data (Using worksheet from handler) ---
        logging.info(f"{prefix} Searching for anchor cell '{anchor_text}'...")
        anchor_cell = _find_anchor_cell(worksheet, anchor_text)
        if not anchor_cell:
            logging.error(f"{prefix} Failed to find anchor cell '{anchor_text}' in sheet '{worksheet.title}'. Load template failed.")
            return False # Close handled in finally

        logging.info(f"{prefix} Anchor found at {anchor_cell.coordinate}. Extracting properties...")
        relative_data = _extract_relative_cells(worksheet, anchor_cell, num_below, num_right)

        if relative_data is not None:
            _template_cache[template_name] = relative_data
            logging.info(f"{prefix} Properties extracted and stored in cache as '{template_name}'.")
            return True # Success, close handled in finally
        else:
            logging.error(f"{prefix} Failed extraction process in sheet '{worksheet.title}' (_extract_relative_cells returned None). Load template failed.")
            return False # Close handled in finally

    except FileNotFoundError as fnf_err: # Catch if handler __init__ fails
        logging.error(f"{prefix} Error initializing ExcelHandler (file not found): {fnf_err}")
        return False
    except Exception as e: # Catch other unexpected errors using handler or during processing
        logging.error(f"{prefix} Unexpected error using ExcelHandler or during template processing: {e}", exc_info=True)
        return False
    finally:
        # Close the handler (which closes the workbook object if open)
        if excel_handler:
            logging.info(f"{prefix} Closing ExcelHandler.")
            excel_handler.close() # Assumes handler has a close method

        # --- Cleanup for the copied file (Keep this part) ---
        if copied_path_created and os.path.exists(copied_path_created):
            try:
                os.remove(copied_path_created)
                logging.info(f"{prefix} Cleaned up temporary copy file: '{copied_path_created}'")
            except Exception as remove_err:
                logging.warning(f"{prefix} Could not remove temporary copy '{copied_path_created}' during cleanup: {remove_err}")


def get_template(template_name: str) -> Optional[Dict[str, Union[Optional[Dict[str, Any]], Optional[List[Optional[Dict[str, Any]]]]]]]:
    """
    Retrieves template data from cache. Returns a shallow copy.
    (Function implementation remains the same as before)
    """
    prefix = f"[TemplateManager.get_template(name='{template_name}')]"
    template = _template_cache.get(template_name)
    if template is not None:
        return template.copy()
    else:
        logging.warning(f"{prefix} Template '{template_name}' not found in cache.") # KEPT WARNING
        return None


def clear_template_cache():
    """Clears all loaded templates from the in-memory cache."""
    prefix = "[TemplateManager.clear_template_cache]"
    global _template_cache
    count = len(_template_cache)
    _template_cache = {}
    logging.info(f"{prefix} Cleared {count} template(s) from the cache.") # Log clear action


def create_xlsx_from_template(
    template_data: Optional[Dict[str, Union[Optional[Dict[str, Any]], Optional[List[Optional[Dict[str, Any]]]]]]],
    output_path: str,
    start_cell_coord: str = "A1",
    sheet_name: str = "Generated Template"
) -> bool:
    """
    Creates XLSX file from template data. Handles anchor merge and places
    vertical lists of cells from 'right_N' keys correctly relative to anchor.
    (Function implementation remains the same as before)
    """
    prefix = "[TemplateManager.create_xlsx_from_template]"

    if not template_data:
        logging.error(f"{prefix} Input template_data is None or empty. Cannot create file.") # KEPT ERROR
        return False

    wb = None
    try:
        wb = Workbook()
        if wb.active:
             ws = wb.active
             ws.title = sheet_name
        else:
             ws = wb.create_sheet(title=sheet_name, index=0)
             if ws is None:
                 logging.error(f"{prefix} Could not get or create active worksheet in new workbook.") # KEPT ERROR
                 return False

        try:
            start_col_str, start_row_str = coordinate_from_string(start_cell_coord)
            start_col_anchor = column_index_from_string(start_col_str)
            start_row_anchor = int(start_row_str)
            if start_col_anchor <= 0 or start_row_anchor <= 0:
                raise ValueError("Start row and column must be positive integers.")
        except (ValueError, TypeError) as e:
            logging.error(f"{prefix} Invalid start_cell_coord '{start_cell_coord}': {e}. Using A1 as default.", exc_info=False) # KEPT ERROR (but use A1)
            start_col_anchor, start_row_anchor = 1, 1

        def _apply_props(target_ws: Worksheet, target_cell: Cell, cell_data_dict: Dict[str, Any], target_r: int, target_c: int):
            """Applies properties from dict to cell. Skips MergedCells."""
            if isinstance(target_cell, MergedCell):
                return # Don't apply props to non-top-left merged cells

            if not isinstance(cell_data_dict, dict):
                logging.warning(f"{prefix} Invalid cell_data_dict type ({type(cell_data_dict)}) for cell {target_cell.coordinate}. Skipping.") # KEPT WARNING
                return

            if 'error' in cell_data_dict:
                logging.warning(f"{prefix} Skipping property application for cell {target_cell.coordinate} due to extraction error: {cell_data_dict['error']}") # KEPT WARNING
                return

            target_cell.value = cell_data_dict.get('value')
            height = cell_data_dict.get('height')
            if height is not None:
                try:
                    height_f = float(height)
                    if height_f >= 0:
                        current_rd = target_ws.row_dimensions.get(target_r)
                        if not current_rd or current_rd.height is None or abs((current_rd.height or 0.0) - height_f) > 1e-6:
                           target_ws.row_dimensions[target_r].height = height_f
                    else:
                        logging.warning(f"{prefix} Ignored negative height {height_f} for row {target_r}") # KEPT WARNING
                except (ValueError, TypeError):
                    logging.warning(f"{prefix} Invalid height value '{height}' for row {target_r}") # KEPT WARNING

            width = cell_data_dict.get('width')
            if width is not None:
                col_letter = get_column_letter(target_c)
                try:
                    width_f = float(width)
                    if width_f >= 0:
                        current_cd = target_ws.column_dimensions.get(col_letter)
                        if not current_cd or current_cd.width is None or abs((current_cd.width or 0.0) - width_f) > 1e-6:
                           target_ws.column_dimensions[col_letter].width = width_f
                    else:
                        logging.warning(f"{prefix} Ignored negative width {width_f} for col {col_letter}") # KEPT WARNING
                except (ValueError, TypeError):
                    logging.warning(f"{prefix} Invalid width value '{width}' for col {col_letter}") # KEPT WARNING

            border_styles = cell_data_dict.get('border')
            if isinstance(border_styles, dict):
                left_side = Side(style=border_styles.get('left')) if border_styles.get('left') else None
                right_side = Side(style=border_styles.get('right')) if border_styles.get('right') else None
                top_side = Side(style=border_styles.get('top')) if border_styles.get('top') else None
                bottom_side = Side(style=border_styles.get('bottom')) if border_styles.get('bottom') else None
                if left_side or right_side or top_side or bottom_side:
                    new_border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)
                    if target_cell.border != new_border:
                         target_cell.border = new_border
                elif target_cell.has_style and target_cell.border != Border():
                     target_cell.border = Border()

            alignment_styles = cell_data_dict.get('alignment')
            if isinstance(alignment_styles, dict):
                h_align = alignment_styles.get('horizontal')
                v_align = alignment_styles.get('vertical')
                w_text = alignment_styles.get('wrap_text')
                if h_align is not None or v_align is not None or w_text is not None:
                    new_alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=w_text)
                    if target_cell.alignment != new_alignment:
                        target_cell.alignment = new_alignment
                elif target_cell.has_style and target_cell.alignment != Alignment():
                     target_cell.alignment = Alignment()

        processed_cells = 0
        max_target_row, max_target_col = 0, 0

        def sort_key_func(k):
            if k == ANCHOR_KEY: return (0, 0)
            parts = k.split('_')
            type_order = 99; index = 0
            if len(parts) == 2 and parts[1].isdigit():
                try:
                    index = int(parts[1])
                    if parts[0] == 'bottom': type_order = 1
                    elif parts[0] == 'right': type_order = 2
                except ValueError: pass
            return (type_order, index)

        valid_template_keys = [k for k in template_data.keys() if not k.startswith('_')]
        sorted_keys = sorted(valid_template_keys, key=sort_key_func)

        for key in sorted_keys:
            cell_value_or_list = template_data[key]
            if cell_value_or_list is None: continue

            if key == ANCHOR_KEY:
                if isinstance(cell_value_or_list, dict) and 'error' not in cell_value_or_list:
                    anchor_data = cell_value_or_list
                    target_row, target_col = start_row_anchor, start_col_anchor
                    if target_row > 0 and target_col > 0:
                        target_cell = ws.cell(row=target_row, column=target_col)
                        _apply_props(ws, target_cell, anchor_data, target_row, target_col)
                        processed_cells += 1
                        max_target_row = max(max_target_row, target_row)
                        max_target_col = max(max_target_col, target_col)
                        merge_info = anchor_data.get('merge_info')
                        if isinstance(merge_info, dict):
                            rows_to_merge = merge_info.get('rows', 1)
                            cols_to_merge = merge_info.get('cols', 1)
                            if rows_to_merge > 1 or cols_to_merge > 1:
                                target_end_row = target_row + rows_to_merge - 1
                                target_end_col = target_col + cols_to_merge - 1
                                try:
                                    merge_range_str = f"{get_column_letter(target_col)}{target_row}:{get_column_letter(target_end_col)}{target_end_row}"
                                    ws.merge_cells(start_row=target_row, start_column=target_col, end_row=target_end_row, end_column=target_end_col)
                                    max_target_row = max(max_target_row, target_end_row)
                                    max_target_col = max(max_target_col, target_end_col)
                                except ValueError as merge_val_err:
                                     logging.error(f"{prefix} Invalid merge range calculated for anchor ({merge_range_str}): {merge_val_err}", exc_info=False)
                                except Exception as merge_err:
                                    logging.error(f"{prefix} Failed to apply merge for anchor ({merge_range_str}): {merge_err}", exc_info=True) # KEPT ERROR
                    else:
                         logging.warning(f"{prefix} Skipping '{key}': Invalid calculated target position R={target_row}, C={target_col}.")
                else:
                    logging.warning(f"{prefix} Skipping key '{key}': Invalid data format ({type(cell_value_or_list)}) or contains error.") # KEPT WARNING
                continue

            elif key.startswith('bottom_'):
                if isinstance(cell_value_or_list, dict) and 'error' not in cell_value_or_list:
                    cell_data_dict = cell_value_or_list
                    try:
                        offset = int(key.split('_')[1])
                        target_row = start_row_anchor + offset
                        target_col = start_col_anchor
                        if target_row <= 0 or target_col <= 0:
                            logging.warning(f"{prefix} Skipping '{key}': Calculated invalid position R={target_row}, C={target_col}.")
                            continue
                        target_cell = ws.cell(row=target_row, column=target_col)
                        _apply_props(ws, target_cell, cell_data_dict, target_row, target_col)
                        if not isinstance(target_cell, MergedCell): processed_cells += 1
                        max_target_row = max(max_target_row, target_row)
                        max_target_col = max(max_target_col, target_col)
                    except (ValueError, IndexError, TypeError) as e:
                        logging.warning(f"{prefix} Skipping '{key}': Error calculating position or applying props: {e}") # KEPT WARNING
                else:
                    logging.warning(f"{prefix} Skipping key '{key}': Invalid data format ({type(cell_value_or_list)}) or contains error.") # KEPT WARNING
                continue

            elif key.startswith('right_'):
                if isinstance(cell_value_or_list, list):
                    cell_data_list = cell_value_or_list
                    try:
                        offset_col = int(key.split('_')[1])
                        target_base_col = start_col_anchor + offset_col
                        if target_base_col <= 0:
                            logging.warning(f"{prefix} Skipping '{key}' list: Calculated invalid base column C={target_base_col}.")
                            continue
                        for row_offset, cell_data_item in enumerate(cell_data_list):
                            if cell_data_item is None: continue
                            if not isinstance(cell_data_item, dict) or 'error' in cell_data_item:
                                logging.warning(f"{prefix}   Skipping row offset {row_offset} for '{key}': Invalid item data ({type(cell_data_item)}).") # KEPT WARNING
                                continue
                            target_row = start_row_anchor + row_offset
                            target_col = target_base_col
                            if target_row <= 0:
                                logging.warning(f"{prefix}   Skipping row offset {row_offset} for '{key}': Calculated invalid row R={target_row}.")
                                continue
                            target_cell = ws.cell(row=target_row, column=target_col)
                            _apply_props(ws, target_cell, cell_data_item, target_row, target_col)
                            if not isinstance(target_cell, MergedCell): processed_cells += 1
                            max_target_row = max(max_target_row, target_row)
                            max_target_col = max(max_target_col, target_col)
                    except (ValueError, IndexError, TypeError) as e:
                        logging.warning(f"{prefix} Error processing '{key}' list: {e}") # KEPT WARNING
                else:
                    logging.warning(f"{prefix} Skipping key '{key}': Expected a list for 'right_N', but got {type(cell_value_or_list)}.") # KEPT WARNING
                continue

            else:
                logging.warning(f"{prefix} Skipping unrecognized key '{key}' in template data.") # KEPT WARNING


        if processed_cells > 0 or (template_data.get(ANCHOR_KEY) and 'error' not in template_data[ANCHOR_KEY]):
            try:
                wb.save(output_path)
                return True
            except Exception as save_err:
                logging.error(f"{prefix} Failed to save workbook to '{output_path}': {save_err}", exc_info=True) # KEPT ERROR
                return False
        else:
            logging.warning(f"{prefix} No valid, non-merged cells processed from template data. Output file '{output_path}' might be empty or only contain a merged anchor.") # KEPT WARNING
            try:
                wb.save(output_path)
                return True
            except Exception as save_err:
                logging.error(f"{prefix} Failed to save potentially empty workbook to '{output_path}': {save_err}", exc_info=True) # KEPT ERROR
                return False

    except Exception as e:
        logging.error(f"{prefix} Failed to create or save workbook at '{output_path}': {e}", exc_info=True) # KEPT ERROR
        return False
    finally:
         if wb:
            try: wb.close()
            except Exception as close_err: logging.warning(f"{prefix} Error closing workbook object during finally block: {close_err}") # KEPT WARNING


# --- Table Insertion Function (from excel_multi_table_creator / template_manager (2)) ---

def add_table_to_sheet(ws, start_row, header_rows, label_data, data_rows, footer_config):
    """
    Inserts a complete table structure (header, labels, data, footer with SUM formulas)
    into an existing worksheet starting at a specific row, shifting existing rows down.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to write to.
        start_row (int): The row number *before* which to insert the table.
        header_rows (list of lists): Data for the header rows.
        label_data (dict): Data for rows identified by labels in the first column.
                           Keys should match the order defined in label_order.
        data_rows (list of lists): The main processed data.
        footer_config (dict): Configuration for the footer.
                                Example: {'keywords': ["TOTAL OF :"], 'calculate_cols': [5, 6, 7, 8, 9], 'static_rows': []}
                                'static_rows' is an optional list of lists for extra rows below the main footer.

    Returns:
        int: The row number immediately after the inserted table (where the next content starts).
             Returns start_row if an error occurs.
    """
    try:
        # --- Calculate total rows needed for the new table ---
        # Define the order of labels to ensure consistent row count
        # IMPORTANT: This label_order is specific to this function's example usage.
        # It might need to be passed as an argument or made more flexible
        # if the labels change between tables added with this function.
        label_order = ["VENDOR#:", "Des : LEATHER", "Case Qty :", "MADE IN CAMBODIA"] # Example order
        num_header_rows = len(header_rows)
        num_label_rows = len(label_order) # Assuming one row per label defined in order
        num_data_rows = len(data_rows)
        # Determine number of footer rows (main calculated row + any static ones)
        num_main_footer_rows = 1 if footer_config.get('keywords') or footer_config.get('calculate_cols') else 0
        num_static_footer_rows = len(footer_config.get('static_rows', []))
        total_table_rows = num_header_rows + num_label_rows + num_data_rows + num_main_footer_rows + num_static_footer_rows

        if total_table_rows <= 0:
             print("Warning: No rows to insert for the table.")
             return start_row

        # --- Insert Blank Rows ---
        print(f"Inserting {total_table_rows} rows at row {start_row}...")
        ws.insert_rows(start_row, amount=total_table_rows)

        # --- Now write into the newly inserted blank rows ---
        current_row = start_row # Start writing at the beginning of the inserted block
        num_header_cols = 0

        # --- Write Header Rows ---
        print(f"Writing header starting at row {current_row}...")
        for row_data in header_rows:
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=cell_value)
            if current_row == start_row:
                num_header_cols = len(row_data)
            # Handle merged cells example
            if "Quantity" in row_data:
                 try:
                     col_index = row_data.index("Quantity") + 1
                     if len(row_data) > col_index and row_data[col_index] is None:
                         ws.merge_cells(start_row=current_row, start_column=col_index, end_row=current_row, end_column=col_index + 1)
                 except ValueError:
                     pass # 'Quantity' not found
            current_row += 1

        # --- Write Label Rows ---
        print(f"Writing label rows starting at row {current_row}...")
        # Use the predefined label_order for consistency
        for label in label_order:
            row_to_write = [None] * num_header_cols
            if label in label_data:
                 provided_data = label_data[label]
                 row_to_write = provided_data[:num_header_cols]
                 row_to_write.extend([None] * (num_header_cols - len(row_to_write)))
            else:
                 row_to_write[0] = label
            for col_idx, cell_value in enumerate(row_to_write, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1

        # --- Write Data Rows ---
        print(f"Writing {len(data_rows)} data rows starting at row {current_row}...")
        table_data_start_row = current_row
        for row_data in data_rows:
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1
        table_data_end_row = current_row - 1

        # --- Write Footer Rows ---
        print(f"Writing footer starting at row {current_row}...")
        # Write the main calculated/keyword footer row (if needed)
        if num_main_footer_rows > 0:
            footer_row_content = [None] * num_header_cols
            footer_keywords = footer_config.get('keywords', [])
            if footer_keywords:
                footer_row_content[0] = footer_keywords[0]

            sum_col_indices = footer_config.get('calculate_cols', [])
            if table_data_start_row <= table_data_end_row:
                for col_index in sum_col_indices:
                    if 1 <= col_index <= num_header_cols:
                        col_letter = get_column_letter(col_index)
                        formula = f"=SUM({col_letter}{table_data_start_row}:{col_letter}{table_data_end_row})"
                        footer_row_content[col_index - 1] = formula
                    else:
                         print(f"Warning: Column index {col_index} for SUM is out of header range (1-{num_header_cols}).")
            else: # No data rows, put 0 in sum columns
                for col_index in sum_col_indices:
                     if 1 <= col_index <= num_header_cols:
                         footer_row_content[col_index - 1] = 0

            # Add other static footer elements (example: item count)
            try:
                # Use first header row for index lookup
                item_no_col_index = header_rows[0].index('ITEM N°') # 0-based index
                footer_row_content[item_no_col_index] = f"{len(data_rows)} ITEMS"
            except (ValueError, IndexError):
                 # Check if 'Product Code' exists as an alternative common identifier
                 try:
                     item_no_col_index = header_rows[0].index('Product Code') # 0-based index
                     footer_row_content[item_no_col_index] = f"{len(data_rows)} ITEMS"
                 except (ValueError, IndexError):
                    print("Warning: Could not determine 'ITEM N°' or 'Product Code' column for item count.")


            for col_idx, cell_value in enumerate(footer_row_content, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1

        # Write any additional static footer rows
        static_footer_rows = footer_config.get('static_rows', [])
        for static_row_data in static_footer_rows:
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1


        print(f"Finished writing table. Next available row is {current_row}")
        # current_row now correctly points to the row after the inserted block
        return current_row

    except Exception as e:
        print(f"An error occurred while inserting table at row {start_row}: {e}")
        logging.error(f"An error occurred while inserting table at row {start_row}: {e}", exc_info=True) # Log with traceback
        # Attempt to return the original start row, although the sheet state might be inconsistent
        return start_row


# --- Example Usage ---
if __name__ == "__main__":
    # Configure basic logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(module)s - %(message)s')

    print("="*30)
    print("--- Running Combined Excel Utilities Example ---")
    print("="*30)
    print(f"Current Working Directory: {os.getcwd()}")

    # --- Example 1: Using Template Manager (from template_manager (1)) ---
    # This part requires the source Excel file and potentially ExcelHandler
    print("\n--- Example 1: Template Manager (Load & Create from Template) ---")
    template_source_workbook = "dap.xlsx" # Make sure this file exists for this example
    template_source_sheet = "Packing list"
    template_anchor = "Mark & Nº"
    template_key = "DapPackingList_v1_Combined"
    template_output_filename = "generated_template_output_combined.xlsx"
    template_start_coord = "C3"

    if not os.path.exists(template_source_workbook):
        logging.warning(f"*** Template source workbook not found at '{os.path.abspath(template_source_workbook)}'. Skipping Example 1. ***")
    elif ExcelHandler is None:
         logging.warning(f"*** ExcelHandler not available. Skipping Example 1 which requires it for loading. ***")
    else:
        # Cleanup previous output
        if os.path.exists(template_output_filename):
            try:
                os.remove(template_output_filename)
                print(f"Removed existing file: '{template_output_filename}'")
            except Exception as e:
                logging.warning(f"Could not remove pre-existing file '{template_output_filename}': {e}")

        print(f"\n>>> Attempting Load Template: Source='{template_source_workbook}', Sheet='{template_source_sheet}', Anchor='{template_anchor}'")
        load_success = load_template(
            workbook_path=template_source_workbook,
            sheet_identifier=template_source_sheet,
            template_name=template_key,
            anchor_text=template_anchor,
            num_below=5, # Example values
            num_right=8, # Example values
            force_reload=True,
            create_copy=False # Set to True if you want to test copy logic
        )

        if load_success:
            print(f"Successfully loaded template as '{template_key}'. Retrieving data...")
            retrieved_data = get_template(template_key)

            if retrieved_data is not None:
                print(f"\n>>> Attempting to create XLSX from template data: Output='{template_output_filename}', Start='{template_start_coord}'")
                create_success = create_xlsx_from_template(
                    template_data=retrieved_data,
                    output_path=template_output_filename,
                    start_cell_coord=template_start_coord,
                    sheet_name=f"Generated_{template_source_sheet}"
                )

                if create_success:
                    print(f"Successfully created output file from template: '{template_output_filename}'")
                    print(f"Location: {os.path.abspath(template_output_filename)}")
                else:
                    logging.error(f"Failed to create output file '{template_output_filename}' from template.")
            else:
                logging.error(f"Failed to retrieve template data ('{template_key}') after load reported success.")
        else:
            logging.error(f"Failed to load template '{template_key}' from '{template_source_workbook}'.")

    # --- Example 2: Using Table Insertion (from excel_multi_table_creator / template_manager (2)) ---
    print("\n--- Example 2: Table Insertion (Insert Multiple Tables) ---")

    # 1. Create a workbook and add some initial content
    wb_insert = openpyxl.Workbook()
    ws_insert = wb_insert.active
    ws_insert.title = "Multiple Reports Inserted"

    ws_insert['A1'] = "EXISTING CONTENT - Line 1"
    ws_insert['A2'] = "EXISTING CONTENT - Line 2"
    ws_insert['A3'] = "EXISTING CONTENT - Line 3"
    ws_insert['A4'] = "This should be pushed down by Table 1"
    ws_insert['A5'] = "This should also be pushed down by Table 1"


    # 2. Define data for the first table to insert
    header1 = [
        ["Mark & N°", "P.O N°", "ITEM N°", "Description", "Quantity", None, "N.W (kgs)", "G.W (kgs)", "CBM"],
        [None, None, None, None, "PCS", "SF", None, None, None]
    ]
    # IMPORTANT: The label keys here MUST match the hardcoded 'label_order' inside add_table_to_sheet
    labels1 = {
        "VENDOR#:": ["VENDOR#:", "9000XYZ", "PO-ABC", None, None, None, None, None, None],
        "Des : LEATHER": ["Des : LEATHER", None, None, "FINISHED LEATHER", None, None, None, None, None],
        "Case Qty :": ["Case Qty :", None, None, None, None, None, None, None, None],
        "MADE IN CAMBODIA": ["MADE IN CAMBODIA", None, None, None, None, None, None, None, None]
    }
    data1 = [
        [None, "9000XYZ", "ITEM001", "Leather Type A", 100, 5000.50, 250.0, 270.0, 1.1],
        [None, "9000XYZ", "ITEM002", "Leather Type B", 150, 7500.00, 350.5, 380.5, 1.5],
    ]
    footer_config1 = {
        'keywords': ["TOTAL OF :"],
        'calculate_cols': [5, 6, 7, 8, 9], # PCS, SF, NW, GW, CBM
        'static_rows': [
            [None, None, None, "HS.CODE: 4107.12.00", None, None, None, None, None]
        ]
    }


    # 3. Define data for the second table to insert
    header2 = [ ["Shipment ID", "Product Code", "Description", "Units", "Value"] ]
    # Pad labels to match the expected 'label_order' in add_table_to_sheet
    # A more robust solution would pass label_order to the function or handle missing keys better.
    labels2_padded = {
        "VENDOR#:": ["Region:", "Asia", None, None, None], # Map Region to VENDOR# position
        "Des : LEATHER": ["Shipped Via:", "Sea", None, None, None], # Map Shipped Via to Des position
        "Case Qty :": ["Case Qty :", None, None, None, None, None, None, None, None], # Placeholder
        "MADE IN CAMBODIA": ["MADE IN CAMBODIA", None, None, None, None, None, None, None, None] # Placeholder
    }
    data2 = [
        ["SHP001", "PROD-X", "Component X", 500, 12500.00],
        ["SHP001", "PROD-Y", "Component Y", 1200, 8400.00],
    ]
    footer_config2 = { 'keywords': ["Subtotal:"], 'calculate_cols': [4, 5] } # Sum Units and Value


    # 4. Add the tables to the sheet sequentially, inserting them
    insert_output_filename = "multi_table_report_inserted_combined.xlsx"
    # Cleanup previous output
    if os.path.exists(insert_output_filename):
        try:
            os.remove(insert_output_filename)
            print(f"Removed existing file: '{insert_output_filename}'")
        except Exception as e:
            logging.warning(f"Could not remove pre-existing file '{insert_output_filename}': {e}")


    # Insert the first table BEFORE row 4 (shifting row 4 and 5 down)
    insert_before_row_1 = 4
    print(f"\n>>> Inserting Table 1 before row {insert_before_row_1} ---")
    next_row_after_table1 = add_table_to_sheet(ws_insert, insert_before_row_1, header1, labels1, data1, footer_config1)

    # Insert the second table AFTER the first table (and the content that was shifted)
    # next_row_after_table1 now holds the row number where the original row 4 content currently resides
    insert_before_row_2 = next_row_after_table1 + 3 # Add some spacing (e.g., 3 rows)
    print(f"\n>>> Inserting Table 2 before row {insert_before_row_2} ---")
    next_row_after_table2 = add_table_to_sheet(ws_insert, insert_before_row_2, header2, labels2_padded, data2, footer_config2)


    # 5. Save the workbook for Example 2
    try:
        wb_insert.save(insert_output_filename)
        print(f"\nSuccessfully created multi-table report with inserts: '{insert_output_filename}'")
        print(f"Location: {os.path.abspath(insert_output_filename)}")
        print(f"Check rows around {insert_before_row_1} and {insert_before_row_2} to see inserted content and shifted original content.")
    except Exception as e:
        print(f"Error saving workbook for Example 2: {e}")
        logging.error(f"Error saving workbook for Example 2: {e}", exc_info=True)
    finally:
        if wb_insert:
             try: wb_insert.close()
             except Exception: pass # Ignore close errors during example


    print("\n--- Combined Example Finished ---")
