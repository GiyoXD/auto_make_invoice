import openpyxl
import os
import math
import json # Added for JSON loading
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet # Optional: for type hinting
from openpyxl.styles import Alignment, Border, Side # Added Border, Side for styling
from typing import List, Dict, Any, Optional, Set # Optional: for type hinting
import traceback # For printing detailed errors

# --- Define Border Style Components ---
# Define a thin side for borders
thin_side = Side(border_style="thin", color="000000")
# Define the standard full grid border (used for columns 2+)
thin_border = Border(left=thin_side,
                     right=thin_side,
                     top=thin_side,
                     bottom=thin_side)

# --- Define Alignment Styles ---
# Default vertical center alignment (used for Vendor row)
vertical_center_alignment = Alignment(vertical='center', wrap_text=True)
# Full center alignment (used for most data/footers)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def add_table_to_sheet(
    ws: Worksheet,
    start_row: int,
    header_rows: List[List[Any]],
    data_rows: List[List[Any]],
    footer_config: Dict[str, Any],
    static_labels: List[str] # Keep parameter for consistency
) -> int:
    """
    Inserts a table structure (header, data, footer with SUM formulas)
    into a worksheet at start_row, shifting existing rows down.
    Merges cells in the header appropriately (vertical and horizontal).
    Applies a specific border style: Col 1 special, Cols 2+ full grid.
    Centers content (except Vendor row which is vertically centered only).
    Places and merges HS Code row correctly.

    Args:
        ws: The openpyxl worksheet object to write to.
        start_row: The row number *before* which to insert the table.
        header_rows: List of lists for header rows (expects 2 rows for proper merging).
        data_rows: List of lists for the main data (including any labels in col 1).
        footer_config: Dict for footer settings.
        static_labels: List of strings used as labels in the first column of initial data rows.

    Returns:
        The row number immediately after the inserted table.
        Returns start_row if an error occurs.
    """

    try:
        # --- Remember the starting row for border application later ---
        table_insertion_start_row = start_row

        # --- Calculate total rows needed ---
        num_header_rows = len(header_rows)
        num_data_rows = len(data_rows)
        num_pre_footer_rows = len(footer_config.get('pre_footer_rows', []))
        num_main_footer_rows = 1 if footer_config.get('keywords') or footer_config.get('calculate_cols') else 0
        num_static_footer_rows = len(footer_config.get('static_rows', []))
        total_table_rows = num_header_rows + num_data_rows + num_pre_footer_rows + num_main_footer_rows + num_static_footer_rows

        if total_table_rows <= 0:
             print("Warning: No rows calculated for the table. Nothing inserted.")
             return start_row

        # --- Insert Blank Rows ---
        ws.insert_rows(start_row, amount=total_table_rows)

        # --- Define Row Ranges and Column Info ---
        current_row = start_row
        num_header_cols = 0
        amount_col_idx = -1 # Initialize Amount column index
        if header_rows and isinstance(header_rows[0], list):
            num_header_cols = len(header_rows[0])
            if num_header_cols == 0: # Handle case of empty header row definition
                print("Warning: Header row defined but has 0 columns.")
                if data_rows: num_header_cols = max((len(r) for r in data_rows if isinstance(r, list)), default=1)
                else: num_header_cols = 1
            # Find "Amount" column index from header
            try:
                amount_col_idx = header_rows[0].index("Amount") + 1
            except ValueError:
                print("Warning: 'Amount' column not found in header for HS Code merging.")


        header_start_row = current_row
        header_end_row = header_start_row + num_header_rows - 1

        table_data_start_row = header_end_row + 1
        table_data_end_row = table_data_start_row + num_data_rows - 1

        pre_footer_start_row = table_data_end_row + 1
        pre_footer_end_row = pre_footer_start_row + num_pre_footer_rows - 1

        main_footer_start_row = pre_footer_end_row + 1
        main_footer_end_row = main_footer_start_row + num_main_footer_rows - 1

        post_footer_start_row = main_footer_end_row + 1
        post_footer_end_row = post_footer_start_row + num_static_footer_rows - 1

        table_end_row = post_footer_end_row # The absolute last row

        # --- Write Header Rows ---
        current_row = header_start_row # Reset current_row for writing
        for row_idx, row_data in enumerate(header_rows):
            actual_row_num = header_start_row + row_idx
            current_row_data = row_data if isinstance(row_data, list) else []
            padded_row_data = current_row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                # Use center_alignment for headers
                ws.cell(row=actual_row_num, column=col_idx, value=cell_value).alignment = center_alignment
            current_row += 1

        # --- Apply Header Merging ---
        # (Merging logic remains the same)
        if num_header_rows == 2 and num_header_cols > 0:
             header_row_1_data = header_rows[0] if isinstance(header_rows[0], list) else []
             header_row_2_data = header_rows[1] if isinstance(header_rows[1], list) else []
             r1, r2 = header_start_row, header_start_row + 1
             for col_idx_0based, val1 in enumerate(header_row_1_data):
                 col_idx_1based = col_idx_0based + 1
                 if col_idx_1based > num_header_cols: continue
                 val2 = header_row_2_data[col_idx_0based] if col_idx_0based < len(header_row_2_data) else None
                 # Horizontal Merge
                 if val1 == "Quantity" and col_idx_0based + 1 < len(header_row_2_data) and header_row_2_data[col_idx_0based + 1] == "SF":
                     try:
                         pcs_col_idx_1based = col_idx_1based; sf_col_idx_1based = col_idx_1based + 1
                         ws.merge_cells(start_row=r1, start_column=pcs_col_idx_1based, end_row=r1, end_column=sf_col_idx_1based)
                         ws.cell(row=r1, column=pcs_col_idx_1based).alignment = center_alignment # Use center
                     except Exception as e: print(f"Warning: H-Merge failed. Error: {e}")
                 # Vertical Merge
                 elif val1 is not None and str(val1).strip() != '' and (val2 is None or str(val2).strip() == ''):
                     is_under_quantity_merge = False
                     try:
                         q_idx = header_row_1_data.index("Quantity")
                         if q_idx <= col_idx_0based <= q_idx + 1 and header_row_2_data[q_idx+1] == "SF": is_under_quantity_merge = True
                     except (ValueError, IndexError): pass
                     if not is_under_quantity_merge:
                         try:
                             ws.merge_cells(start_row=r1, start_column=col_idx_1based, end_row=r2, end_column=col_idx_1based)
                             ws.cell(row=r1, column=col_idx_1based).alignment = center_alignment # Use center
                         except Exception as e: print(f"Error V-Merging Col {col_idx_1based}: {e}")


        # --- Write Data Rows ---
        current_row = table_data_start_row # Start writing data
        for i, row_data in enumerate(data_rows):
            actual_row_num = table_data_start_row + i
            current_row_data = row_data if isinstance(row_data, list) else []
            padded_row_data = current_row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 processed_value = cell_value
                 try: # Numeric conversion
                     if isinstance(cell_value, str) and cell_value and (cell_value.replace('.', '', 1).isdigit() or (cell_value.startswith('-') and cell_value[1:].replace('.', '', 1).isdigit())):
                          numeric_val = float(cell_value)
                          processed_value = int(numeric_val) if numeric_val.is_integer() else numeric_val
                 except (ValueError, TypeError): pass
                 cell = ws.cell(row=actual_row_num, column=col_idx, value=processed_value)
                 # Apply alignment conditionally based on row index 'i'
                 if i == 0 and col_idx == 1: # First data row (VENDOR#:) Column 1
                     cell.alignment = vertical_center_alignment # Only vertical center
                 else: # Other data rows OR other columns in first row
                     cell.alignment = center_alignment # Full center
            current_row += 1

        # --- Write Pre-Footer Static Rows (Placeholders) ---
        current_row = pre_footer_start_row # Start writing pre-footer
        # This loop writes the placeholder structure (e.g., [None, None, ...])
        for i, static_row_data in enumerate(footer_config.get('pre_footer_rows', [])):
            actual_row_num = pre_footer_start_row + i
            current_row_data = static_row_data if isinstance(static_row_data, list) else []
            padded_row_data = current_row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 # Write the placeholder value (likely None) and center it
                 ws.cell(row=actual_row_num, column=col_idx, value=cell_value).alignment = center_alignment
            current_row += 1

        # --- Apply HS Code Merging AFTER writing the placeholder row ---
        # This assumes the first pre_footer_row is intended for HS Code
        if num_pre_footer_rows > 0:
             hs_code_row_num = pre_footer_start_row # Row number for HS Code
             print(f"Applying HS Code merge logic to row {hs_code_row_num}")
             hs_code_text = "HS.CODE: 4107.XX.XX" # Define the text

             # Write HS Code text to Col 2, overwriting placeholder
             cell_hs = ws.cell(row=hs_code_row_num, column=2, value=hs_code_text)
             cell_hs.alignment = center_alignment # Center align HS code block

             # Merge Col 2 and 3 for HS Code
             try:
                 ws.merge_cells(start_row=hs_code_row_num, start_column=2, end_row=hs_code_row_num, end_column=3)
                 print(f"Merged HS Code text in row {hs_code_row_num} across columns 2-3")
             except Exception as merge_err:
                 print(f"Error merging HS Code text (cols 2-3) in row {hs_code_row_num}: {merge_err}")

             # Merge empty cells to the right (Col 4 up to Amount col - 1)
             if amount_col_idx > 4: # Check if Amount col exists and there's space to merge
                 merge_start_col = 4
                 merge_end_col = amount_col_idx - 1
                 try:
                     ws.merge_cells(start_row=hs_code_row_num, start_column=merge_start_col, end_row=hs_code_row_num, end_column=merge_end_col)
                     print(f"Merged empty cells in row {hs_code_row_num} from column {merge_start_col} to {merge_end_col}")
                     # Ensure the top-left cell of this merge is blank and centered
                     ws.cell(row=hs_code_row_num, column=merge_start_col, value=None).alignment = center_alignment
                 except Exception as merge_err:
                     print(f"Error merging empty cells (cols {merge_start_col}-{merge_end_col}) in row {hs_code_row_num}: {merge_err}")
             elif amount_col_idx != -1:
                  print(f"Not enough columns between HS Code and Amount (Col {amount_col_idx}) to merge empty cells.")

             # Ensure Col 1 is blank and centered in HS Code row (overwriting placeholder)
             ws.cell(row=hs_code_row_num, column=1, value=None).alignment = center_alignment


        # --- Write Main Footer Row (Totals) ---
        current_row = main_footer_start_row # Start writing main footer
        if num_main_footer_rows > 0:
            actual_row_num = main_footer_start_row
            # Summation
            sum_col_indices = footer_config.get('calculate_cols', [])
            if table_data_start_row <= table_data_end_row: # Check if data rows exist
                for col_index in sum_col_indices:
                    if 1 <= col_index <= num_header_cols:
                        col_letter = get_column_letter(col_index)
                        formula = f"=SUM({col_letter}{table_data_start_row}:{col_letter}{table_data_end_row})"
                        # Apply center alignment to total cells
                        ws.cell(row=actual_row_num, column=col_index, value=formula).alignment = center_alignment
            else: # No data rows
                for col_index in sum_col_indices:
                     if 1 <= col_index <= num_header_cols:
                         # Apply center alignment
                         ws.cell(row=actual_row_num, column=col_index, value=0).alignment = center_alignment
            # Item count
            if header_rows and isinstance(header_rows[0], list):
                try:
                    item_col_name = next((name for name in ['ITEM N°', 'Product Code'] if name in header_rows[0]), None)
                    if item_col_name:
                        item_no_col_index = header_rows[0].index(item_col_name) + 1
                        if 1 <= item_no_col_index <= num_header_cols:
                            # Apply center alignment
                            ws.cell(row=actual_row_num, column=item_no_col_index, value=f"{len(data_rows)} ITEMS").alignment = center_alignment
                except (ValueError, IndexError): pass
            # Keyword
            footer_keywords = footer_config.get('keywords', [])
            if footer_keywords and ws.cell(row=actual_row_num, column=1).value is None:
                # Apply center alignment
                ws.cell(row=actual_row_num, column=1, value=footer_keywords[0]).alignment = center_alignment

            # Ensure remaining cells in footer row are centered if they get values later
            for c_idx in range(1, num_header_cols + 1):
                if ws.cell(row=actual_row_num, column=c_idx).value is None:
                     ws.cell(row=actual_row_num, column=c_idx).alignment = center_alignment

            current_row += 1

        # --- Write Post-Footer Static Rows ---
        current_row = post_footer_start_row # Start writing post-footer
        for i, static_row_data in enumerate(footer_config.get('static_rows', [])):
            actual_row_num = post_footer_start_row + i
            current_row_data = static_row_data if isinstance(static_row_data, list) else []
            padded_row_data = current_row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 # Apply center alignment
                 ws.cell(row=actual_row_num, column=col_idx, value=cell_value).alignment = center_alignment
            current_row += 1


        # --- Apply Borders Based on Column and Section ---
        # (Border logic remains the same column-specific logic)
        if num_header_cols > 0:
            print(f"Applying column-specific borders from row {table_insertion_start_row} to {table_end_row}, col 1 to {num_header_cols}")

            for r in range(table_insertion_start_row, table_end_row + 1):
                for c in range(1, num_header_cols + 1):
                    cell = ws.cell(row=r, column=c)

                    # --- Logic for Columns 2+ (Normal Full Grid) ---
                    if c > 1:
                        cell.border = thin_border
                        continue # Move to next cell

                    # --- Logic for Column 1 Only ---
                    left_side = thin_side # Always Left
                    right_side = thin_side # Always Right (separator to col 2)
                    top_side = None
                    bottom_side = None

                    # Determine Top/Bottom based on row's section for Col 1
                    if header_start_row <= r <= header_end_row: # Header
                        top_side = thin_side; bottom_side = thin_side
                    elif table_data_start_row <= r <= table_data_end_row: # Data
                        if r == table_data_start_row: top_side = thin_side
                        if r == table_data_end_row: bottom_side = thin_side
                    elif pre_footer_start_row <= r <= pre_footer_end_row: # Pre-footer
                        top_side = thin_side; bottom_side = thin_side
                    elif main_footer_start_row <= r <= main_footer_end_row: # Main Footer
                        top_side = thin_side; bottom_side = thin_side
                    elif post_footer_start_row <= r <= post_footer_end_row: # Post-footer
                        top_side = thin_side; bottom_side = thin_side

                    cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)
        else:
            print("Skipping border application as num_header_cols is 0.")


        # Return the row number *after* the last written row
        return table_end_row + 1

    except Exception as e:
        print(f"--- ERROR occurred while inserting table starting near row {start_row} ---")
        print(f"Error details: {e}")
        traceback.print_exc()
        # Try to return a safe row number
        return current_row if 'current_row' in locals() else start_row + 1


# --- Main Execution Logic ---
if __name__ == "__main__":

    json_input_filename = "test.json"
    # Output name reflects latest changes
    output_filename = "tables_from_json_final_style.xlsx"

    # --- 1. Load JSON Data ---
    if not os.path.exists(json_input_filename):
        print(f"Error: JSON input file '{json_input_filename}' not found."); exit()
    try:
        with open(json_input_filename, 'r', encoding='utf-8') as f: json_data = json.load(f)
        print(f"Successfully loaded data from '{json_input_filename}'")
    except Exception as e: print(f"Error loading JSON: {e}"); traceback.print_exc(); exit()

    # --- 2. Prepare Workbook ---
    if os.path.exists(output_filename):
        try: os.remove(output_filename); print(f"Removed previous output file: '{output_filename}'")
        except Exception as e: print(f"Warning: Could not remove file '{output_filename}': {e}")

    try:
        wb = openpyxl.Workbook()
        sheet_name = json_data.get('metadata', {}).get('worksheet_name', 'Inserted Tables Report')
        sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_')).rstrip()[:31]
        ws = wb.active; ws.title = sheet_name
        print(f"Created workbook with sheet: '{ws.title}'")
    except Exception as e: print(f"Error creating workbook: {e}"); traceback.print_exc(); exit()

    ws['A1'] = f"Report: {json_data.get('metadata', {}).get('workbook_filename', 'JSON Data')}"
    ws['A2'] = f"Sheet: {ws.title}"; ws['A3'] = ""

    # --- 3. Process and Insert Tables from JSON ---
    next_row_to_insert = 4

    if 'processed_tables_data' not in json_data or not isinstance(json_data['processed_tables_data'], dict):
         print("Error: JSON data missing 'processed_tables_data' dictionary."); exit()

    static_labels_in_data = ["VENDOR#:", "Des : LEATHER", "Case Qty :", "MADE IN CAMBODIA"]

    for table_id, table_data in json_data['processed_tables_data'].items():
        print(f"\n--- Processing Table ID: {table_id} ---")
        if not isinstance(table_data, dict) or not table_data:
            print(f"Warning: Skipping Table ID '{table_id}' - invalid/empty data."); continue

        # --- 3a. Define Header & Map JSON keys ---
        header_rows = [
            ["Mark & N°", "P.O N°", "ITEM N°", "Description", "Quantity", None, "N.W (kgs)", "G.W (kgs)", "CBM", "Unit", "Amount"],
            [None, None, None, None, "PCS", "SF", None, None, None, None, None]
        ]
        col_map = {
            "po": 2, "item": 3, "reference_code": 4, "pcs": 5, "sqft": 6,
            "net": 7, "gross": 8, "cbm": 9, "unit": 10, "amount": 11
        }
        num_cols = len(header_rows[0]) if header_rows and isinstance(header_rows[0], list) else 0
        if num_cols == 0: continue # Skip if no columns

        mapped_json_keys = list(col_map.keys())
        missing_keys = [k for k in mapped_json_keys if k not in table_data]
        if missing_keys:
             print(f"Warning: Skipping Table ID '{table_id}'. Missing keys: {missing_keys}"); continue

        # --- 3b. Prepare data_rows ---
        data_rows = []
        try:
            max_data_len = 0
            for key in mapped_json_keys:
                if key in table_data and isinstance(table_data[key], list):
                    max_data_len = max(max_data_len, len(table_data[key]))
            num_data_entries = max_data_len
            num_label_rows = len(static_labels_in_data)
            total_data_section_rows = max(num_data_entries, num_label_rows)

            for i in range(total_data_section_rows):
                row = [None] * num_cols
                if i < num_label_rows: row[0] = static_labels_in_data[i]
                for key, col_idx in col_map.items():
                    if 1 <= col_idx <= num_cols:
                        if key in table_data and isinstance(table_data[key], list) and i < len(table_data[key]):
                            row[col_idx - 1] = table_data[key][i]
                data_rows.append(row)
        except Exception as e: print(f"Error data_rows prep: {e}"); traceback.print_exc(); continue

        # --- 3c. Define Footer Configuration ---
        # Define hs_code_row structure with None values initially
        hs_code_row_structure = [None] * num_cols

        footer_config = {
            'keywords': [f"TOTALS (Table {table_id}):"],
            'calculate_cols': [c for c in [col_map.get(k) for k in ["pcs", "sqft", "net", "gross", "cbm", "amount"]] if c is not None],
             'pre_footer_rows': [ hs_code_row_structure ], # Pass the row structure
            'static_rows': [ [None] * num_cols ] # Blank row after totals
        }

        # --- 3d. Call add_table_to_sheet ---
        print(f"Inserting table for ID '{table_id}' before row {next_row_to_insert}...")
        try:
            next_row_after_insertion = add_table_to_sheet(
                ws,
                next_row_to_insert,
                header_rows,
                data_rows,
                footer_config,
                static_labels_in_data # Pass labels
            )
            print(f"Table ID '{table_id}' finished. Next content starts at row {next_row_after_insertion}")
            next_row_to_insert = next_row_after_insertion + 1 # Add spacing row
        except Exception as insert_error:
             print(f"--- FAILED to insert Table ID '{table_id}' ---"); print(f"Error: {insert_error}"); traceback.print_exc()
             next_row_to_insert += 1

    # --- 4. Auto-adjust column widths ---
    print("\n--- Adjusting column widths ---")
    column_widths = {}
    for row_idx in range(1, ws.max_row + 1):
         max_col_check = max(ws.max_column, 50) # Check reasonable number of columns
         for col_idx in range(1, max_col_check + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    # Check if cell is part of a merged range
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                             # Use the top-left cell of the merge for width calculation
                             if cell.coordinate == merged_range.min_row_col_coord:
                                  cell_length = len(str(cell.value))
                                  column_widths[col_idx] = max(column_widths.get(col_idx, 0), cell_length)
                             is_merged = True
                             break
                    if not is_merged:
                        cell_length = len(str(cell.value))
                        column_widths[col_idx] = max(column_widths.get(col_idx, 0), cell_length)
                except Exception: pass # Ignore errors

    for col_idx, max_length in column_widths.items():
        column_letter = get_column_letter(col_idx)
        # Add slightly more padding
        adjusted_width = min(max_length + 3, 70) # Increased padding and max width
        ws.column_dimensions[column_letter].width = adjusted_width


    # --- 5. Save the final workbook ---
    try:
        wb.save(output_filename)
        print(f"\n--- Workbook saved successfully: '{output_filename}' ---")
        print(f"Full path: {os.path.abspath(output_filename)}")
    except PermissionError:
         print(f"\n--- ERROR saving workbook: Permission denied. Is '{output_filename}' open? ---")
         traceback.print_exc()
    except Exception as e:
        print(f"\n--- ERROR saving workbook: {e} ---"); traceback.print_exc()
    finally:
        if 'wb' in locals() and wb:
             try: wb.close()
             except Exception: pass

    print("\n--- Script finished ---")
