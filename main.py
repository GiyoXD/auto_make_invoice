import openpyxl
import os
import math
import json # Added for JSON loading
import re # Added for regex in find_all_header_rows
import logging # Added for logging in find_all_header_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet # Optional: for type hinting
# Updated typing import to include List, Dict, Any, Optional
from typing import List, Dict, Any, Optional, Set # Added Set for map_columns_to_headers
import traceback # For printing detailed errors

# Configure basic logging (optional, but good practice if using logging)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Configuration Constants (PLACEHOLDERS - Configure these!) ---
# TARGET_HEADERS_MAP: Maps canonical names to lists of possible variations found in Excel.
# Example: {'po': ['P.O N°', 'PO Number', 'Purchase Order'], 'item': ['ITEM N°', 'Item Code'], ...}
TARGET_HEADERS_MAP: Dict[str, List[str]] = {
    # Add your canonical header names and their possible variations here
    # Example:
    # 'po': ['P.O N°', 'PO Number', 'Purchase Order'],
    # 'item': ['ITEM N°', 'Item Code', 'Product ID'],
    # 'reference_code': ['Description', 'Ref Code'],
    # 'pcs': ['Quantity', 'PCS', 'Qty'],
    # 'sqft': ['SF', 'SQFT', 'Area'],
    # 'net': ['N.W (kgs)', 'Net Weight'],
    # 'gross': ['G.W (kgs)', 'Gross Weight'],
    # 'cbm': ['CBM', 'Volume'],
    # 'unit': ['Unit', 'Unit Price'],
    # 'amount': ['Amount', 'Total Value']
}

# DISTRIBUTION_BASIS_COLUMN: Canonical name of the column used as the basis for distribution (e.g., 'sqft').
DISTRIBUTION_BASIS_COLUMN: Optional[str] = None # e.g., 'sqft'

# COLUMNS_TO_DISTRIBUTE: List of canonical names of columns whose values should be distributed.
COLUMNS_TO_DISTRIBUTE: List[str] = [] # e.g., ['net', 'gross', 'cbm']
# --- End Configuration Constants ---


def find_all_header_rows(sheet: Worksheet, search_pattern: str, row_range: int, col_range: int) -> List[int]:
    """
    Finds all 1-indexed row numbers containing a header based on a pattern.
    Returns a list of row numbers, sorted in ascending order.

    Args:
        sheet: The openpyxl worksheet object to search in.
        search_pattern: The regex pattern to search for in cell values.
        row_range: The maximum row number to search up to.
        col_range: The maximum column number to search up to.

    Returns:
        A list of 1-based row indices where the header pattern was found, sorted.
        Returns an empty list if no match is found or an error occurs.
    """
    header_rows: List[int] = []
    try:
        # Compile the regex pattern once, ignore case
        regex = re.compile(search_pattern, re.IGNORECASE)

        # Determine search boundaries, ensuring they don't exceed sheet dimensions
        # Use sheet.max_row and sheet.max_column for accurate limits
        max_row_to_search = min(row_range, sheet.max_row)
        max_col_to_search = min(col_range, sheet.max_column)

        logging.info(f"[find_all_header_rows] Searching for headers using pattern '{search_pattern}' in rows 1-{max_row_to_search}, cols 1-{max_col_to_search}")

        # Iterate through the specified range to find header cells
        for r_idx in range(1, max_row_to_search + 1):
            # Optimization: Check only necessary columns if pattern is specific (can be enhanced)
            for c_idx in range(1, max_col_to_search + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                # Check if cell has a value before trying to convert to string
                if cell.value is not None:
                    # Convert cell value to string and strip whitespace
                    cell_value_str = str(cell.value).strip()
                    # If the cell content matches the pattern, consider this a header row
                    if regex.search(cell_value_str):
                        logging.debug(f"[find_all_header_rows] Header pattern found in cell {cell.coordinate} (Row: {r_idx}). Adding row to list.")
                        # Check if this row is already added to prevent duplicates
                        # if multiple cells match in the same row
                        if r_idx not in header_rows:
                            header_rows.append(r_idx)
                        # Optimization: Once a header is found in a row,
                        # no need to check other columns in the same row. Move to the next row.
                        break # Break inner column loop

        # Sort the found header rows in ascending order
        header_rows.sort()

        if not header_rows:
            # Use logging.warning for non-critical findings like 'not found'
            logging.warning(f"[find_all_header_rows] Header pattern '{search_pattern}' not found within the search range (Rows 1-{max_row_to_search}, Cols 1-{max_col_to_search}).")
        else:
            # Log the final list found AT INFO level for clarity
            logging.info(f"[find_all_header_rows] Found {len(header_rows)} potential header rows at: {header_rows}")

        return header_rows
    except re.error as regex_err:
        # Catch specific regex compilation errors
        logging.error(f"[find_all_header_rows] Invalid regex pattern '{search_pattern}': {regex_err}", exc_info=False)
        return []
    except Exception as e:
        # Catch any other unexpected errors during the process
        logging.error(f"[find_all_header_rows] Error finding header rows: {e}", exc_info=True)
        return []


def map_columns_to_headers(sheet: Worksheet, header_row: int, col_range: int) -> Dict[str, int]:
    """
    Maps canonical header names to their 1-indexed column numbers based on the
    header row content, prioritizing the first match found based on TARGET_HEADERS_MAP order.
    (Uses the variation -> canonical lookup for clarity)

    Args:
        sheet: The openpyxl worksheet object.
        header_row: The 1-indexed row number containing the headers.
        col_range: The maximum number of columns to search for headers.

    Returns:
        A dictionary mapping canonical names (str) to column indices (int).
        Returns an empty dictionary if the header row is invalid or no mapping is found.
    """
    if header_row is None or header_row < 1:
        logging.error("[map_columns_to_headers] Invalid header_row provided for column mapping: %s", header_row)
        return {}
    # Ensure TARGET_HEADERS_MAP is defined and is a dictionary
    if not isinstance(TARGET_HEADERS_MAP, dict):
        logging.error("[map_columns_to_headers] Configuration Error: TARGET_HEADERS_MAP is not defined or not a dictionary.")
        return {}


    column_mapping: Dict[str, int] = {}
    processed_canonicals: Set[str] = set() # Track canonical names already assigned to a column
    max_col_to_check = min(col_range, sheet.max_column)

    logging.info(f"[map_columns_to_headers] Mapping columns based on header row {header_row} up to column {max_col_to_check}.")

    # --- Build a reverse lookup: lowercase variation -> canonical name ---
    variation_to_canonical_lookup: Dict[str, str] = {}
    ambiguous_variations: Set[str] = set()
    for canonical_name, variations in TARGET_HEADERS_MAP.items():
        # Ensure variations is iterable (list or tuple), even if it's a single string in the config
        actual_variations: List[str] = []
        if isinstance(variations, str):
            actual_variations = [variations] # Treat single string as a list with one item
        elif isinstance(variations, (list, tuple)):
             actual_variations = list(variations) # Use list or tuple directly
        else:
             logging.error(f"[map_columns_to_headers] Config Error: Value for canonical name '{canonical_name}' in TARGET_HEADERS_MAP is not a list, tuple, or string: {variations}. Skipping this canonical name.")
             continue # Skip this canonical name if variations format is wrong

        # Process each variation for the current canonical name
        for variation in actual_variations:
            # Ensure variation is treated as string and handle potential None values
            variation_str = str(variation) if variation is not None else ""
            variation_lower = variation_str.lower().strip()
            if not variation_lower: continue # Skip empty variations

            # Check for ambiguity: if variation already maps to a *different* canonical name
            if variation_lower in variation_to_canonical_lookup and variation_to_canonical_lookup[variation_lower] != canonical_name:
                 # Log ambiguity only once per variation
                 if variation_lower not in ambiguous_variations:
                      logging.warning(f"[map_columns_to_headers] Config Ambiguity: Header variation '{variation_lower}' is mapped to multiple canonical names ('{variation_to_canonical_lookup[variation_lower]}' and '{canonical_name}'). Using the first one encountered.")
                      ambiguous_variations.add(variation_lower)
                 # Decision: Keep the first mapping encountered based on TARGET_HEADERS_MAP iteration order. Do not overwrite.
            elif variation_lower not in variation_to_canonical_lookup:
                 # Add the mapping if the variation is not already mapped
                 variation_to_canonical_lookup[variation_lower] = canonical_name

    # --- Iterate through Excel columns and map using the lookup ---
    logging.debug(f"[map_columns_to_headers] Built variation lookup: {variation_to_canonical_lookup}")
    for col_idx in range(1, max_col_to_check + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        # Use .value directly; openpyxl handles data types. Convert to string for lookup.
        cell_value = cell.value
        # Convert to string, handle None, lowercase, and strip whitespace
        actual_header_text = str(cell_value).lower().strip() if cell_value is not None else ""

        if not actual_header_text:
            # Log empty header cells at DEBUG level
            logging.debug(f"[map_columns_to_headers] Cell {get_column_letter(col_idx)}{header_row} in header row {header_row} is empty or None.")
            continue

        # Find the canonical name corresponding to the actual header text found in the sheet
        matched_canonical = variation_to_canonical_lookup.get(actual_header_text)

        if matched_canonical:
            # Check if this canonical name has already been assigned to a column
            if matched_canonical not in processed_canonicals:
                # Assign the column index to the canonical name
                column_mapping[matched_canonical] = col_idx
                # Mark this canonical name as processed
                processed_canonicals.add(matched_canonical)
                # Log successful mapping at INFO level for clarity
                logging.info(f"[map_columns_to_headers] Mapped column {col_idx} (Header Text: '{cell.value}') -> Canonical: '{matched_canonical}'")
            else:
                # Log if we find another column with a header that maps to an already assigned canonical name
                original_col = column_mapping.get(matched_canonical, 'UNKNOWN') # Should exist, but safety check
                logging.warning(f"[map_columns_to_headers] Duplicate Canonical Mapping Ignored: Header '{cell.value}' in Col {col_idx} also maps to canonical '{matched_canonical}', which was already assigned to Col {original_col}. Keeping first assignment.")
        else:
             # Log headers found in Excel but not matching any known variation at DEBUG level
             # This helps identify headers in the sheet that might need to be added to TARGET_HEADERS_MAP
             logging.debug(f"[map_columns_to_headers] Excel header '{cell.value}' (Col {col_idx}) in row {header_row} did not match any known variations in TARGET_HEADERS_MAP.")


    # --- Post-mapping checks ---
    if not column_mapping:
        logging.warning(f"[map_columns_to_headers] No target headers were successfully mapped in row {header_row} up to column {max_col_to_check}. Check Excel headers and TARGET_HEADERS_MAP content.")
    else:
        logging.info(f"[map_columns_to_headers] Completed mapping for header row {header_row}. Found mappings: {column_mapping}")
        # Check for essential columns needed later (using the global config constants)
        required_columns: Set[str] = set()
        # Add distribution basis column if defined
        if DISTRIBUTION_BASIS_COLUMN and isinstance(DISTRIBUTION_BASIS_COLUMN, str):
            required_columns.add(DISTRIBUTION_BASIS_COLUMN)
        # Add columns to distribute if defined
        if COLUMNS_TO_DISTRIBUTE and isinstance(COLUMNS_TO_DISTRIBUTE, list):
            required_columns.update(COLUMNS_TO_DISTRIBUTE)
        # Also check essentials for SQFT aggregation if known (example)
        # Adjust these based on actual downstream needs
        required_columns.update(['po', 'item', 'unit', 'sqft']) # Example essential columns

        # Find which required columns are missing from the actual mapping results
        missing_essentials = required_columns - set(column_mapping.keys())
        if missing_essentials:
             # This is important, log as WARNING
             logging.warning(f"[map_columns_to_headers] Mapping complete for row {header_row}, but MISSING essential canonical mappings needed for processing: {sorted(list(missing_essentials))}. Subsequent steps might fail.")
        else:
             # Log success if all checked essential columns were found
             logging.info(f"[map_columns_to_headers] All checked essential columns ({sorted(list(required_columns))}) appear to be mapped successfully for header row {header_row}.")


    return column_mapping


def add_table_to_sheet(
    ws: Worksheet,
    start_row: int,
    header_rows: List[List[Any]],
    # label_data removed - labels are now part of data_rows
    data_rows: List[List[Any]],
    footer_config: Dict[str, Any]
    # label_order removed
) -> int:
    """
    Inserts a table structure (header, data, footer with SUM formulas)
    into a worksheet at start_row, shifting existing rows down.
    Assumes labels like 'VENDOR#:' are included in the first column of data_rows.

    Args:
        ws: The openpyxl worksheet object to write to.
        start_row: The row number *before* which to insert the table.
        header_rows: List of lists for header rows.
        data_rows: List of lists for the main data (including any labels in col 1).
        footer_config: Dict for footer settings. Example:
                       {'keywords': ["TOTAL:"], 'calculate_cols': [5, 6],
                        'pre_footer_rows': [[None,'Comment',None]], # Static rows *before* totals
                        'static_rows': [[None, 'Comment', None]]} # Static rows *after* totals

    Returns:
        The row number immediately after the inserted table.
        Returns start_row if an error occurs.
    """
    try:
        # --- Calculate total rows needed ---
        num_header_rows = len(header_rows)
        num_data_rows = len(data_rows)
        # Re-introduced pre_footer_rows calculation
        num_pre_footer_rows = len(footer_config.get('pre_footer_rows', [])) # Rows *before* totals
        num_main_footer_rows = 1 if footer_config.get('keywords') or footer_config.get('calculate_cols') else 0
        num_static_footer_rows = len(footer_config.get('static_rows', [])) # Rows *after* totals
        # Adjusted total calculation
        total_table_rows = num_header_rows + num_data_rows + num_pre_footer_rows + num_main_footer_rows + num_static_footer_rows

        if total_table_rows <= 0:
             logging.warning(f"[add_table_to_sheet] No rows calculated for the table at start_row {start_row}. Nothing inserted.")
             return start_row

        # --- Insert Blank Rows ---
        logging.info(f"[add_table_to_sheet] Inserting {total_table_rows} blank rows before row {start_row}.")
        ws.insert_rows(start_row, amount=total_table_rows)

        # --- Write into the newly inserted blank rows ---
        current_row = start_row
        num_header_cols = 0
        if header_rows:
            # Determine number of columns from the first header row
            num_header_cols = len(header_rows[0])
            logging.debug(f"[add_table_to_sheet] Determined {num_header_cols} columns based on header.")
        elif data_rows:
            # Fallback: Determine from first data row if no header
            num_header_cols = len(data_rows[0])
            logging.warning(f"[add_table_to_sheet] No header rows provided. Determined {num_header_cols} columns from first data row.")
        else:
            logging.warning("[add_table_to_sheet] No header or data rows provided. Cannot determine column count.")
            # Decide how to handle this - maybe return? For now, proceed cautiously.
            # num_header_cols remains 0, loops below might not execute correctly.

        # --- Write Header Rows ---
        logging.debug(f"[add_table_to_sheet] Writing {num_header_rows} header rows starting at row {current_row}.")
        for row_idx, row_data in enumerate(header_rows):
            # Pad row data to match the determined number of columns
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=cell_value)
            # Handle merged cells example (specific to the structure in the example)
            if "Quantity" in row_data:
                 try:
                     # Find the index of "Quantity"
                     q_idx = row_data.index("Quantity")
                     # Check if the next cell exists and is intended for merging (e.g., is None)
                     if q_idx + 1 < len(row_data) and row_data[q_idx + 1] is None:
                         col_index_to_merge = q_idx + 1 # 1-based index for openpyxl
                         # Merge "Quantity" cell with the one to its right
                         ws.merge_cells(start_row=current_row, start_column=col_index_to_merge, end_row=current_row, end_column=col_index_to_merge + 1)
                         logging.debug(f"[add_table_to_sheet] Merged cells {get_column_letter(col_index_to_merge)}{current_row}:{get_column_letter(col_index_to_merge+1)}{current_row}")
                 except ValueError:
                     # "Quantity" not found in this specific header row
                     pass
                 except Exception as merge_err:
                     logging.error(f"[add_table_to_sheet] Error merging cells at row {current_row}: {merge_err}", exc_info=False)
            current_row += 1

        # --- Write Data Rows (now starts immediately after header) ---
        table_data_start_row = current_row
        logging.debug(f"[add_table_to_sheet] Writing {num_data_rows} data rows starting at row {current_row}.")
        for row_data in data_rows:
            # Pad row data consistently
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 # Attempt numeric conversion for potential string numbers
                 try:
                     # Check if it's a string that looks like a number (int or float)
                     if isinstance(cell_value, str) and cell_value.strip():
                         # Handle negative numbers and decimals
                         cleaned_value = cell_value.strip()
                         if (cleaned_value.replace('.', '', 1).isdigit() or
                             (cleaned_value.startswith('-') and cleaned_value[1:].replace('.', '', 1).isdigit())):
                              cell_value_numeric = float(cleaned_value)
                              # Convert to int if it's a whole number
                              cell_value = int(cell_value_numeric) if cell_value_numeric.is_integer() else cell_value_numeric
                 except (ValueError, TypeError):
                     # If conversion fails, keep the original value
                     pass
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1
        table_data_end_row = current_row - 1
        logging.debug(f"[add_table_to_sheet] Data rows written from {table_data_start_row} to {table_data_end_row}.")

        # --- Write Pre-Footer Static Rows (e.g., HS CODE) ---
        pre_footer_rows = footer_config.get('pre_footer_rows', [])
        logging.debug(f"[add_table_to_sheet] Writing {len(pre_footer_rows)} pre-footer rows starting at row {current_row}.")
        for static_row_data in pre_footer_rows:
             # Pad row data
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        # --- Write Main Footer Row (Totals) ---
        if num_main_footer_rows > 0:
            logging.debug(f"[add_table_to_sheet] Writing main footer row (totals) at row {current_row}.")
            # Initialize footer row content with Nones
            footer_row_content = [None] * num_header_cols
            footer_keywords = footer_config.get('keywords', [])
            if footer_keywords:
                # Place the first keyword in the first column (index 0)
                footer_row_content[0] = footer_keywords[0]

            # Add SUM formulas
            sum_col_indices = footer_config.get('calculate_cols', [])
            # Check if there are actual data rows to sum
            if table_data_start_row <= table_data_end_row:
                for col_index in sum_col_indices:
                    # Ensure column index is valid (1-based)
                    if 1 <= col_index <= num_header_cols:
                        col_letter = get_column_letter(col_index)
                        formula = f"=SUM({col_letter}{table_data_start_row}:{col_letter}{table_data_end_row})"
                        # Place formula in the correct 0-based index of the list
                        footer_row_content[col_index - 1] = formula
                        logging.debug(f"[add_table_to_sheet] Added SUM formula '{formula}' to footer column {col_index}.")
                    else:
                         logging.warning(f"[add_table_to_sheet] Footer SUM column index {col_index} is out of range (1-{num_header_cols}). Skipping.")
            else: # No data rows were added, set sums to 0
                logging.warning(f"[add_table_to_sheet] No data rows found between {table_data_start_row} and {table_data_end_row}. Setting SUM columns to 0.")
                for col_index in sum_col_indices:
                     if 1 <= col_index <= num_header_cols:
                         footer_row_content[col_index - 1] = 0 # Set to 0 if no data

            # Add item count (example logic, adjust as needed)
            # Assumes header_rows[0] contains the definitive header names
            if header_rows and data_rows: # Only add count if there's a header and data
                try:
                    # Find the column name for items (case-insensitive check might be better)
                    item_col_name = next((name for name in ['ITEM N°', 'Product Code', 'Item No'] if name in header_rows[0]), None)
                    if item_col_name:
                        # Find the 0-based index of that column name in the first header row
                        item_no_col_index_0based = header_rows[0].index(item_col_name)
                        # Add the count string to the footer row content at that index
                        footer_row_content[item_no_col_index_0based] = f"{len(data_rows)} ITEMS"
                        logging.debug(f"[add_table_to_sheet] Added item count '{len(data_rows)} ITEMS' to footer column {item_no_col_index_0based + 1}.")
                    else:
                        logging.warning("[add_table_to_sheet] Could not find a recognized item column ('ITEM N°', 'Product Code', 'Item No') in header to add item count.")
                except (ValueError, IndexError, StopIteration) as count_err:
                     # Catch potential errors during index finding or iteration
                     logging.warning(f"[add_table_to_sheet] Could not determine item column for item count: {count_err}")

            # Write the constructed footer row
            for col_idx, cell_value in enumerate(footer_row_content, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1

        # --- Write Post-Footer Static Rows ---
        static_footer_rows = footer_config.get('static_rows', [])
        logging.debug(f"[add_table_to_sheet] Writing {len(static_footer_rows)} post-footer rows starting at row {current_row}.")
        for static_row_data in static_footer_rows:
             # Pad row data
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        logging.info(f"[add_table_to_sheet] Successfully inserted table. Next available row is {current_row}.")
        return current_row # Return the next row index *after* the entire table

    except Exception as e:
        # Log the error with traceback for detailed debugging
        logging.error(f"--- ERROR occurred while inserting table intended to start at row {start_row} ---", exc_info=True)
        # Optionally print to console as well
        print(f"--- ERROR occurred while inserting table at row {start_row} ---")
        print(f"Error details: {e}")
        traceback.print_exc()
        # Return the original start_row to indicate failure or minimal progress
        return start_row


# --- Main Execution Logic ---
if __name__ == "__main__":

    json_input_filename = "test.json"
    output_filename = "tables_from_json_corrected.xlsx"

    # --- 1. Load JSON Data ---
    if not os.path.exists(json_input_filename):
        logging.error(f"JSON input file '{json_input_filename}' not found.")
        print(f"Error: JSON input file '{json_input_filename}' not found.") # Keep console output
        exit()
    try:
        with open(json_input_filename, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        logging.info(f"Successfully loaded data from '{json_input_filename}'")
        print(f"Successfully loaded data from '{json_input_filename}'")
    except json.JSONDecodeError as json_err:
        logging.error(f"Error decoding JSON file '{json_input_filename}': {json_err}", exc_info=True)
        print(f"Error loading or parsing JSON file '{json_input_filename}': {json_err}")
        traceback.print_exc(); exit()
    except Exception as e:
        logging.error(f"Unexpected error loading JSON file '{json_input_filename}': {e}", exc_info=True)
        print(f"Error loading or parsing JSON file '{json_input_filename}': {e}")
        traceback.print_exc(); exit()

    # --- 2. Prepare Workbook ---
    if os.path.exists(output_filename):
        try:
            os.remove(output_filename)
            logging.info(f"Removed previous output file: '{output_filename}'")
            print(f"Removed previous output file: '{output_filename}'")
        except Exception as e:
            logging.warning(f"Could not remove existing file '{output_filename}': {e}")
            print(f"Warning: Could not remove existing file '{output_filename}': {e}")

    try:
        wb = openpyxl.Workbook()
        # Sanitize sheet name (max 31 chars, no invalid chars)
        sheet_name_raw = json_data.get('metadata', {}).get('worksheet_name', 'Inserted Tables Report')
        # Remove invalid characters: []:*?/\\
        sheet_name_sanitized = re.sub(r'[\[\]:*?/\\]', '', sheet_name_raw)
        # Truncate to 31 characters
        sheet_name = sheet_name_sanitized[:31]
        ws = wb.active
        ws.title = sheet_name
        logging.info(f"Created workbook with sheet: '{ws.title}'")
        print(f"Created workbook with sheet: '{ws.title}'")
    except Exception as e:
        logging.error(f"Error creating workbook: {e}", exc_info=True)
        print(f"Error creating workbook: {e}"); traceback.print_exc(); exit()

    # Add initial content (optional report header)
    ws['A1'] = f"Report Source: {json_data.get('metadata', {}).get('workbook_filename', json_input_filename)}"
    ws['A2'] = f"Worksheet: {ws.title}"
    ws['A3'] = "" # Empty row for spacing

    # --- 3. Process and Insert Tables from JSON ---
    next_row_to_insert = 4 # Start inserting after initial content

    processed_tables = json_data.get('processed_tables_data')
    if not isinstance(processed_tables, dict):
         logging.error("JSON data missing 'processed_tables_data' dictionary or it's not a dictionary.")
         print("Error: JSON data missing 'processed_tables_data' dictionary or it's not a dictionary."); exit()

    # Define the static labels that should appear in the first column of the first few data rows
    # These seem specific to the example data structure
    static_labels_in_data = ["VENDOR#:", "Des : LEATHER", "Case Qty :", "MADE IN CAMBODIA"]
    logging.info(f"Static labels defined for data rows: {static_labels_in_data}")

    # Iterate through tables found in the JSON data
    for table_id, table_data in processed_tables.items():
        logging.info(f"\n--- Processing Table ID: {table_id} ---")
        print(f"\n--- Processing Table ID: {table_id} ---") # Keep console output

        if not isinstance(table_data, dict) or not table_data:
            logging.warning(f"Skipping Table ID '{table_id}' due to invalid/empty data.")
            print(f"Warning: Skipping Table ID '{table_id}' due to invalid/empty data."); continue

        # --- 3a. Define Header & Map JSON keys ---
        # This structure seems specific to the expected input JSON
        header_rows = [
            ["Mark & N°", "P.O N°", "ITEM N°", "Description", "Quantity", None, "N.W (kgs)", "G.W (kgs)", "CBM", "Unit", "Amount"],
            [None, None, None, None, "PCS", "SF", None, None, None, None, None] # Second header row for units/details
        ]
        # Map JSON keys (lowercase) to 1-based column indices in the header
        # NOTE: This col_map is specific to the *output* structure defined in header_rows.
        # It's NOT derived from reading the input Excel file's headers.
        # The map_columns_to_headers function (if used before this) would handle input mapping.
        col_map = {
            "po": 2, "item": 3, "reference_code": 4, "pcs": 5, "sqft": 6,
            "net": 7, "gross": 8, "cbm": 9, "unit": 10, "amount": 11
            # "mark_no" could potentially map to column 1 if needed
        }
        num_cols = len(header_rows[0]) if header_rows else 0
        logging.debug(f"Table '{table_id}': Output header defined with {num_cols} columns. Output column map: {col_map}")

        # --- 3b. Prepare data_rows (Integrate static labels) ---
        data_rows = []
        try:
            # Find the length of the first list in the table_data dictionary to determine number of rows
            # Assumes all lists under table_data[key] have the same length for this table
            first_key = next(iter(table_data))
            if not isinstance(table_data[first_key], list):
                 raise TypeError(f"Expected list for key '{first_key}' in table '{table_id}', got {type(table_data[first_key])}")
            num_rows_in_table = len(table_data[first_key])
            logging.debug(f"Table '{table_id}': Found {num_rows_in_table} data rows based on key '{first_key}'.")

            # Check if all expected keys (based on col_map for output) exist in this table's JSON data
            expected_keys = list(col_map.keys())
            missing_keys = [k for k in expected_keys if k not in table_data]
            if missing_keys:
                 # Log warning but proceed, assuming missing data will result in None/empty cells
                 logging.warning(f"Table ID '{table_id}': Missing expected data keys in JSON: {missing_keys}. Corresponding output columns will be empty.")
                 # Decide if this should be fatal: print(...) ; continue

            # Check if all corresponding values are lists (for keys that *are* present)
            non_list_keys = [k for k in expected_keys if k in table_data and not isinstance(table_data.get(k), list)]
            if non_list_keys:
                logging.warning(f"Skipping Table ID '{table_id}'. Keys with non-list data: {non_list_keys}")
                print(f"Warning: Skipping Table ID '{table_id}'. Keys with non-list data: {non_list_keys}"); continue

            # Construct each data row for the output table
            for i in range(num_rows_in_table):
                # Initialize row with None values for all output columns
                row = [None] * num_cols
                # Set the first column label if applicable for this row index
                if i < len(static_labels_in_data):
                    row[0] = static_labels_in_data[i]
                # else: row[0] remains None

                # Populate the rest of the row from JSON data using col_map
                for key, col_idx_1based in col_map.items():
                    # Ensure column index is within bounds of the output table
                    if 1 <= col_idx_1based <= num_cols:
                        col_idx_0based = col_idx_1based - 1
                        # Check if data exists for this key *in the JSON* at this index `i`
                        if key in table_data and i < len(table_data[key]):
                             row[col_idx_0based] = table_data[key][i]
                        else:
                             # Log if data is missing or JSON list is shorter than expected
                             if key in table_data: # Key exists, but list is too short
                                 logging.warning(f"Data missing for key '{key}' at index {i} in Table '{table_id}'. Setting cell to None.")
                             # else: key was missing entirely (already warned above)
                             row[col_idx_0based] = None # Explicitly set to None
                    # else: col_idx_1based is out of bounds (shouldn't happen if col_map matches header_rows)

                data_rows.append(row)
            logging.debug(f"Table '{table_id}': Prepared {len(data_rows)} data rows for output.")

        except StopIteration:
            logging.warning(f"Table '{table_id}' appears to have no data keys in JSON. Skipping.");
            print(f"Warning: Table '{table_id}' empty. Skipping."); continue
        except TypeError as te:
            logging.error(f"Type error processing data for Table '{table_id}': {te}. Skipping.", exc_info=False)
            print(f"Error transforming data for Table '{table_id}': {te}. Skipping."); continue
        except Exception as e:
            logging.error(f"Unexpected error transforming data for Table '{table_id}': {e}", exc_info=True)
            print(f"Error transforming data for Table '{table_id}': {e}"); traceback.print_exc(); continue

        # --- 3c. Define Footer Configuration ---
        # Define the HS Code row to be inserted *before* the totals
        # Ensure it has the correct number of columns based on the output header
        hs_code_row = [None] * num_cols
        if num_cols >= 4: # Check if there are enough columns for the HS code placement
            hs_code_row[3] = "HS.CODE: 4107.XX.XX" # Place in 4th column (index 3)
        else:
            logging.warning(f"Table '{table_id}': Not enough output columns ({num_cols}) to place HS Code in the 4th column.")
            # Decide fallback: maybe place in last available column or omit?
            if num_cols > 0: hs_code_row[0] = "HS.CODE: 4107.XX.XX" # Example: place in first col if fewer than 4

        footer_config = {
            'keywords': [f"TOTALS (Table {table_id}):"], # Keyword for the first column of the totals row
            'calculate_cols': [ # List of 1-based column indices to SUM (using the output col_map)
                col_map['pcs'], col_map['sqft'], col_map['net'],
                col_map['gross'], col_map['cbm'], col_map['amount']
             ],
             'pre_footer_rows': [ hs_code_row ], # List of rows to insert *before* the main totals row
            'static_rows': [ # List of rows to insert *after* the main totals row
                 [None] * num_cols, # Add an empty row for spacing after totals
                 ]
        }
        logging.debug(f"Table '{table_id}': Footer configuration defined: {footer_config}")

        # --- 3d. Call add_table_to_sheet ---
        logging.info(f"Inserting table for ID '{table_id}' before row {next_row_to_insert}...")
        print(f"Inserting table for ID '{table_id}' before row {next_row_to_insert}...")
        try:
            # Call the function to insert the table into the worksheet
            next_row_after_insertion = add_table_to_sheet(
                ws=ws,
                start_row=next_row_to_insert,
                header_rows=header_rows,
                data_rows=data_rows,
                footer_config=footer_config
            )
            # Check if insertion was successful (returned row > start_row)
            if next_row_after_insertion > next_row_to_insert:
                logging.info(f"Table ID '{table_id}' inserted successfully. Next content would start at row {next_row_after_insertion}")
                print(f"Table ID '{table_id}' finished. Next content would start at row {next_row_after_insertion}")
                # Update the next insertion point, adding 1 for spacing between tables
                next_row_to_insert = next_row_after_insertion + 1
            else:
                # Insertion likely failed, log warning and increment minimally
                logging.warning(f"Insertion of Table ID '{table_id}' might have failed (returned row {next_row_after_insertion} <= start row {next_row_to_insert}). Incrementing row minimally.")
                print(f"Warning: Insertion of Table ID '{table_id}' might have failed.")
                next_row_to_insert += 1 # Increment minimally to avoid potential infinite loops

        except Exception as insert_error:
             # Catch any unexpected errors during the add_table_to_sheet call
             logging.error(f"--- FAILED to insert Table ID '{table_id}' due to unexpected error: {insert_error} ---", exc_info=True)
             print(f"--- FAILED to insert Table ID '{table_id}' ---"); print(f"Error: {insert_error}"); traceback.print_exc()
             next_row_to_insert += 1 # Increment minimally to try the next table

    # --- 4. Save the final workbook ---
    try:
        wb.save(output_filename)
        abs_path = os.path.abspath(output_filename)
        logging.info(f"--- Workbook saved successfully: '{output_filename}' ---")
        logging.info(f"Full path: {abs_path}")
        print(f"\n--- Workbook saved successfully: '{output_filename}' ---")
        print(f"Full path: {abs_path}")
    except PermissionError as pe:
         logging.error(f"--- ERROR saving workbook: Permission denied. File '{output_filename}' might be open in another program. {pe} ---", exc_info=False)
         print(f"\n--- ERROR saving workbook: Permission denied. Please close '{output_filename}' if it's open and try again. ---")
    except Exception as e:
        logging.error(f"--- ERROR saving workbook: {e} ---", exc_info=True);
        print(f"\n--- ERROR saving workbook: {e} ---"); traceback.print_exc()
    finally:
        # Ensure workbook is closed even if saving fails
        if 'wb' in locals() and wb:
             try:
                 wb.close()
                 logging.debug("Workbook closed.")
             except Exception as close_err:
                 # This error during close is less critical but good to know
                 logging.warning(f"Error closing workbook object: {close_err}", exc_info=False)

    logging.info("--- Script finished ---")
    print("\n--- Script finished ---")
