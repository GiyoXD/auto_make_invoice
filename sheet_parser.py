import openpyxl
import os
import math
import json # Added for JSON loading
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet # Optional: for type hinting
from typing import List, Dict, Any, Optional # Optional: for type hinting
import traceback # For printing detailed errors

def add_table_to_sheet(
    ws: Worksheet,
    start_row: int,
    header_rows: List[List[Any]],
    # label_data removed - labels are now part of data_rows
    data_rows: List[List[Any]],
    footer_config: Dict[str, Any]
    # label_order removed
) -> int:
    """
    Inserts a table structure (header, data, footer with SUM formulas)
    into a worksheet at start_row, shifting existing rows down.
    Assumes labels like 'VENDOR#:' are included in the first column of data_rows.

    Args:
        ws: The openpyxl worksheet object to write to.
        start_row: The row number *before* which to insert the table.
        header_rows: List of lists for header rows.
        data_rows: List of lists for the main data (including any labels in col 1).
        footer_config: Dict for footer settings. Example:
                       {'keywords': ["TOTAL:"], 'calculate_cols': [5, 6],
                        'pre_footer_rows': [[None,'Comment',None]], # Static rows *before* totals
                        'static_rows': [[None, 'Comment', None]]} # Static rows *after* totals

    Returns:
        The row number immediately after the inserted table.
        Returns start_row if an error occurs.
    """
    try:
        # --- Calculate total rows needed ---
        num_header_rows = len(header_rows)
        num_data_rows = len(data_rows)
        # Re-introduced pre_footer_rows calculation
        num_pre_footer_rows = len(footer_config.get('pre_footer_rows', [])) # Rows *before* totals
        num_main_footer_rows = 1 if footer_config.get('keywords') or footer_config.get('calculate_cols') else 0
        num_static_footer_rows = len(footer_config.get('static_rows', [])) # Rows *after* totals
        # Adjusted total calculation
        total_table_rows = num_header_rows + num_data_rows + num_pre_footer_rows + num_main_footer_rows + num_static_footer_rows

        if total_table_rows <= 0:
             print("Warning: No rows calculated for the table. Nothing inserted.")
             return start_row

        # --- Insert Blank Rows ---
        ws.insert_rows(start_row, amount=total_table_rows)

        # --- Write into the newly inserted blank rows ---
        current_row = start_row
        num_header_cols = 0
        if header_rows:
            num_header_cols = len(header_rows[0])

        # --- Write Header Rows ---
        for row_data in header_rows:
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=cell_value)
            # Handle merged cells example
            if "Quantity" in row_data:
                 try:
                     q_idx = row_data.index("Quantity")
                     if q_idx + 1 < len(row_data) and row_data[q_idx + 1] is None:
                         col_index = q_idx + 1
                         ws.merge_cells(start_row=current_row, start_column=col_index, end_row=current_row, end_column=col_index + 1)
                 except ValueError: pass
            current_row += 1

        # --- Write Data Rows (now starts immediately after header) ---
        table_data_start_row = current_row
        for row_data in data_rows:
            padded_row_data = row_data[:num_header_cols]
            padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
            for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 # Attempt numeric conversion
                 try:
                     if isinstance(cell_value, str) and (cell_value.replace('.', '', 1).isdigit() or (cell_value.startswith('-') and cell_value[1:].replace('.', '', 1).isdigit())):
                          cell_value_numeric = float(cell_value)
                          cell_value = int(cell_value_numeric) if cell_value_numeric.is_integer() else cell_value_numeric
                 except (ValueError, TypeError): pass
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1
        table_data_end_row = current_row - 1

        # --- Write Pre-Footer Static Rows (e.g., HS CODE) ---
        # This section is re-introduced
        pre_footer_rows = footer_config.get('pre_footer_rows', [])
        for static_row_data in pre_footer_rows:
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        # --- Write Main Footer Row (Totals) ---
        if num_main_footer_rows > 0:
            footer_row_content = [None] * num_header_cols
            footer_keywords = footer_config.get('keywords', [])
            if footer_keywords:
                footer_row_content[0] = footer_keywords[0]

            # Add SUM formulas
            sum_col_indices = footer_config.get('calculate_cols', [])
            if table_data_start_row <= table_data_end_row:
                for col_index in sum_col_indices:
                    if 1 <= col_index <= num_header_cols:
                        col_letter = get_column_letter(col_index)
                        formula = f"=SUM({col_letter}{table_data_start_row}:{col_letter}{table_data_end_row})"
                        footer_row_content[col_index - 1] = formula
                    else:
                         print(f"Warning: Footer SUM col index {col_index} out of range (1-{num_header_cols}).")
            else: # No data rows
                for col_index in sum_col_indices:
                     if 1 <= col_index <= num_header_cols: footer_row_content[col_index - 1] = 0

            # Add item count
            if header_rows:
                try:
                    item_col_name = next((name for name in ['ITEM N°', 'Product Code'] if name in header_rows[0]), None)
                    if item_col_name:
                        item_no_col_index = header_rows[0].index(item_col_name)
                        footer_row_content[item_no_col_index] = f"{len(data_rows)} ITEMS"
                except (ValueError, IndexError):
                     print("Warning: Could not determine item column for item count.")

            # Write the footer row
            for col_idx, cell_value in enumerate(footer_row_content, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1

        # --- Write Post-Footer Static Rows ---
        # These rows appear *after* the main totals row
        static_footer_rows = footer_config.get('static_rows', [])
        for static_row_data in static_footer_rows:
             padded_row_data = static_row_data[:num_header_cols]
             padded_row_data.extend([None] * (num_header_cols - len(padded_row_data)))
             for col_idx, cell_value in enumerate(padded_row_data, start=1):
                 ws.cell(row=current_row, column=col_idx, value=cell_value)
             current_row += 1

        return current_row

    except Exception as e:
        print(f"--- ERROR occurred while inserting table at row {start_row} ---")
        print(f"Error details: {e}")
        traceback.print_exc()
        return start_row


# --- Main Execution Logic ---
if __name__ == "__main__":

    json_input_filename = "test.json"
    output_filename = "tables_from_json_corrected.xlsx"

    # --- 1. Load JSON Data ---
    if not os.path.exists(json_input_filename):
        print(f"Error: JSON input file '{json_input_filename}' not found.")
        exit()
    try:
        with open(json_input_filename, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        print(f"Successfully loaded data from '{json_input_filename}'")
    except Exception as e:
        print(f"Error loading or parsing JSON file '{json_input_filename}': {e}")
        traceback.print_exc(); exit()

    # --- 2. Prepare Workbook ---
    if os.path.exists(output_filename):
        try: os.remove(output_filename); print(f"Removed previous output file: '{output_filename}'")
        except Exception as e: print(f"Warning: Could not remove existing file '{output_filename}': {e}")

    try:
        wb = openpyxl.Workbook()
        sheet_name = json_data.get('metadata', {}).get('worksheet_name', 'Inserted Tables Report')
        sheet_name = sheet_name[:31].replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace(':', '').replace('\\', '').replace('/', '')
        ws = wb.active; ws.title = sheet_name
        print(f"Created workbook with sheet: '{ws.title}'")
    except Exception as e: print(f"Error creating workbook: {e}"); traceback.print_exc(); exit()

    # Add initial content (optional)
    ws['A1'] = f"Report: {json_data.get('metadata', {}).get('workbook_filename', 'JSON Data')}"
    ws['A2'] = f"Sheet: {ws.title}"; ws['A3'] = ""

    # --- 3. Process and Insert Tables from JSON ---
    next_row_to_insert = 4 # Start inserting after initial content

    if 'processed_tables_data' not in json_data or not isinstance(json_data['processed_tables_data'], dict):
         print("Error: JSON data missing 'processed_tables_data' dictionary."); exit()

    # Define the static labels that should appear in the first column of the first few data rows
    static_labels_in_data = ["VENDOR#:", "Des : LEATHER", "Case Qty :", "MADE IN CAMBODIA"]

    for table_id, table_data in json_data['processed_tables_data'].items():
        print(f"\n--- Processing Table ID: {table_id} ---")
        if not isinstance(table_data, dict) or not table_data:
            print(f"Warning: Skipping Table ID '{table_id}' due to invalid/empty data."); continue

        # --- 3a. Define Header & Map JSON keys ---
        header_rows = [
            ["Mark & N°", "P.O N°", "ITEM N°", "Description", "Quantity", None, "N.W (kgs)", "G.W (kgs)", "CBM", "Unit", "Amount"],
            [None, None, None, None, "PCS", "SF", None, None, None, None, None]
        ]
        col_map = {
            "po": 2, "item": 3, "reference_code": 4, "pcs": 5, "sqft": 6,
            "net": 7, "gross": 8, "cbm": 9, "unit": 10, "amount": 11
        }
        expected_keys = list(col_map.keys())
        if not all(key in table_data for key in expected_keys):
             missing = [k for k in expected_keys if k not in table_data]
             print(f"Warning: Skipping Table ID '{table_id}'. Missing keys: {missing}"); continue
        num_cols = len(header_rows[0]) if header_rows else 0

        # --- 3b. Prepare data_rows (Integrate static labels) ---
        data_rows = []
        try:
            first_key = next(iter(table_data))
            num_rows = len(table_data[first_key])

            for i in range(num_rows):
                row = [None] * num_cols # Initialize row
                # Set the first column label if applicable for this row index
                if i < len(static_labels_in_data):
                    row[0] = static_labels_in_data[i]
                else:
                    row[0] = None # No static label for this data row

                # Populate the rest of the row from JSON data using col_map
                for key, col_idx in col_map.items():
                    if col_idx > num_cols: continue
                    if i < len(table_data[key]):
                         row[col_idx - 1] = table_data[key][i]
                    else:
                         print(f"Warning: Data missing for key '{key}' at index {i} in Table '{table_id}'.")
                         row[col_idx - 1] = None
                data_rows.append(row)

        except StopIteration: print(f"Warning: Table '{table_id}' empty. Skipping."); continue
        except Exception as e: print(f"Error transforming data for Table '{table_id}': {e}"); traceback.print_exc(); continue

        # --- 3c. Define Footer Configuration ---
        # Define the HS Code row to be inserted *before* the totals
        hs_code_row = [None, None, None, "HS.CODE: 4107.XX.XX"] + [None]*(num_cols-4) # Adjust column index if needed

        footer_config = {
            'keywords': [f"TOTALS (Table {table_id}):"],
            'calculate_cols': [
                col_map['pcs'], col_map['sqft'], col_map['net'],
                col_map['gross'], col_map['cbm'], col_map['amount']
             ],
             'pre_footer_rows': [ hs_code_row ], # HS Code row before totals
            'static_rows': [ # Optional rows *after* the totals
                 [None] * num_cols, # Add empty row after totals
                 ]
        }

        # --- 3d. Call add_table_to_sheet ---
        print(f"Inserting table for ID '{table_id}' before row {next_row_to_insert}...")
        try:
            # Call the function
            next_row_after_insertion = add_table_to_sheet(
                ws,
                next_row_to_insert,
                header_rows,
                data_rows,
                footer_config
            )
            print(f"Table ID '{table_id}' finished. Next content would start at row {next_row_after_insertion}")
            next_row_to_insert = next_row_after_insertion + 1 # Add spacing

        except Exception as insert_error:
             print(f"--- FAILED to insert Table ID '{table_id}' ---"); print(f"Error: {insert_error}"); traceback.print_exc()
             next_row_to_insert += 1 # Increment minimally

    # --- 4. Save the final workbook ---
    try:
        wb.save(output_filename)
        print(f"\n--- Workbook saved successfully: '{output_filename}' ---")
        print(f"Full path: {os.path.abspath(output_filename)}")
    except Exception as e:
        print(f"\n--- ERROR saving workbook: {e} ---"); traceback.print_exc()
    finally:
        if wb:
             try: wb.close()
             except Exception: pass

    print("\n--- Script finished ---")
