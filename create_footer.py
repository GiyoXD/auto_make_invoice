# create_footer.py
# Loads extracted row properties from JSON, finds a row marker and a column marker
# in a target Excel file, overwrites the row starting AT the column marker.
# Borders are applied to the overwritten row based on the maximum column width
# found in reference rows (marked by "Mark & Nº" or "NO." in Col A).

import logging
import os
import json
import pprint
import argparse # For command-line arguments
import re # For column marker pattern matching
from typing import Dict, List, Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font

# --- Script Constants ---
# Default values if not provided via arguments
DEFAULT_INPUT_JSON = "footer.json" # JSON from total_row_property_extracter.py
DEFAULT_ROW_MARKER_TEXT = "footer1" # Marker to find the ROW to overwrite
DEFAULT_COL_MARKER_PATTERN = r"footer1" # Regex pattern to find START COLUMN (e.g., matches FOOTER1, FOOTER2 etc.)
# Optional: Limit the search for markers to the first N columns
MARKER_SEARCH_MAX_COL = 10 # Search range for row and column markers
# Default dimensions if not specified in JSON
DEFAULT_COL_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0
# Text markers for reference rows used to determine max border column
REFERENCE_ROW_MARKERS = ["Mark & Nº", "NO."] # Case-insensitive check in Column A

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Define Styles ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
no_border = Border()
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def find_marker_row(worksheet: Worksheet, marker_text: str, max_col_to_search: int) -> Optional[int]:
    """
    Finds the 1-based index of the first row containing the marker_text
    within the first few columns.
    """
    logging.debug(f"Searching for row marker '{marker_text}' in sheet '{worksheet.title}'...")
    max_col = min(max_col_to_search, worksheet.max_column)
    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    # Ensure comparison is done with string representation
                    if marker_text == str(cell.value).strip():
                        logging.info(f"Found row marker '{marker_text}' in cell {cell.coordinate} (Row: {row_idx}).")
                        return row_idx
                except Exception as e:
                    logging.warning(f"Could not check cell {cell.coordinate} value '{cell.value}': {e}")
                    continue
    logging.warning(f"Row marker text '{marker_text}' not found in sheet '{worksheet.title}' within first {max_col} columns.")
    return None

def find_start_column(worksheet: Worksheet, target_row_index: int, col_pattern: str, max_col_to_search: int) -> int:
    """
    Finds the 1-based index of the first column in the target_row
    containing text matching the col_pattern. Defaults to 1 if not found.
    """
    logging.debug(f"Searching for column pattern '{col_pattern}' in row {target_row_index}...")
    try:
        # Compile regex for case-insensitive matching
        regex = re.compile(col_pattern, re.IGNORECASE)
    except re.error as e:
        logging.error(f"Invalid column regex pattern '{col_pattern}': {e}. Defaulting to column 1.")
        return 1 # Default to column 1 on pattern error

    max_col = min(max_col_to_search, worksheet.max_column)
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=target_row_index, column=col_idx)
        if cell.value is not None:
            try:
                # Use regex search to find pattern anywhere in the cell value
                if regex.search(str(cell.value).strip()):
                    logging.info(f"Found column marker pattern '{col_pattern}' in cell {cell.coordinate} (Column: {col_idx}). Using as start column.")
                    return col_idx
            except Exception as e:
                 logging.warning(f"Could not check cell {cell.coordinate} value '{cell.value}': {e}")
                 continue
    logging.warning(f"Column marker pattern '{col_pattern}' not found in row {target_row_index} within first {max_col} columns. Defaulting to start column 1.")
    return 1 # Default to column 1 if no marker found

def find_last_col_with_value(worksheet: Worksheet, row_idx: int) -> int:
    """
    Finds the 1-based index of the last column in a given row that contains a value.
    Returns 0 if the row is empty or not found.
    """
    last_col = 0
    # Iterate from max column backwards to find the last cell with content
    for col_idx in range(worksheet.max_column, 0, -1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if cell.value is not None and str(cell.value).strip() != "":
            last_col = col_idx
            break # Found the last one
    if last_col > 0:
        logging.debug(f"Found last column with value in row {row_idx}: {last_col}")
    else:
        logging.debug(f"Row {row_idx} appears empty or has no values.")
    return last_col

def apply_cell_properties(ws: Worksheet, target_cell: openpyxl.cell.Cell, cell_props: Dict[str, Any], apply_border: bool):
    """
    Applies properties from a dictionary (value, width, height, border, etc.)
    to a specific target cell. Border application is controlled by apply_border flag.
    """
    try:
        # Apply Value
        value_to_apply = cell_props.get('value')
        target_cell.value = value_to_apply

        # Apply Alignment
        target_cell.alignment = center_alignment

        # --- Conditional Border Application ---
        if apply_border:
            # Apply the simple thin border if requested
            target_cell.border = thin_border
        else:
            # Ensure no border if flag is False
             if target_cell.has_style and target_cell.border != no_border:
                 target_cell.border = no_border
        # --- END Conditional Border ---

        # Apply Column Width
        width = cell_props.get('width')
        col_letter = target_cell.column_letter
        if width is not None:
            try:
                width_float = float(width)
                current_width = ws.column_dimensions[col_letter].width
                if current_width is None or abs(current_width - width_float) > 1e-6:
                    ws.column_dimensions[col_letter].width = width_float
            except (ValueError, TypeError):
                logging.warning(f"Invalid width '{width}' for {target_cell.coordinate}, using default.")
                if ws.column_dimensions[col_letter].width is None: ws.column_dimensions[col_letter].width = DEFAULT_COL_WIDTH
        elif ws.column_dimensions[col_letter].width is None:
             ws.column_dimensions[col_letter].width = DEFAULT_COL_WIDTH

    except Exception as e:
        logging.error(f"Error applying properties to cell {target_cell.coordinate}: {e}", exc_info=True)

def overwrite_row_with_footer(ws: Worksheet, target_row_index: int, start_column_index: int, footer_cell_data: List[Dict[str, Any]], max_border_col_index: int):
    """
    OVERWRITES cells in the target row, starting AT start_column_index,
    with data and properties from footer_cell_data.
    Applies borders only up to max_border_col_index (determined from reference rows).

    Args:
        ws: The target worksheet.
        target_row_index: The row number to overwrite.
        start_column_index: The starting column index (1-based) where the marker was found.
        footer_cell_data: List of dictionaries, each containing properties for one cell.
        max_border_col_index: The maximum column index to apply borders up to.
    """
    if not footer_cell_data:
        logging.warning("No footer cell data provided. Cannot overwrite row.")
        return

    num_footer_cells = len(footer_cell_data)
    logging.info(f"Overwriting row {target_row_index} starting AT column {start_column_index} with {num_footer_cells} cells in sheet '{ws.title}'...")
    if max_border_col_index > 0:
        logging.info(f"Borders will be applied up to column {max_border_col_index} based on reference rows.")
    else:
        logging.warning("No valid reference rows found or they were empty. Borders will not be applied to the footer.")

    try:
        # Apply row height using the height from the first cell's properties
        if footer_cell_data:
            height = footer_cell_data[0].get('height')
            if height is not None:
                try:
                    ws.row_dimensions[target_row_index].height = float(height)
                except (ValueError, TypeError):
                     logging.warning(f"Invalid height '{height}' for target row {target_row_index}. Using default.")
                     ws.row_dimensions[target_row_index].height = DEFAULT_ROW_HEIGHT
            else:
                 ws.row_dimensions[target_row_index].height = DEFAULT_ROW_HEIGHT
        else:
             ws.row_dimensions[target_row_index].height = DEFAULT_ROW_HEIGHT


        # Populate the target row starting from start_column_index
        max_col_in_sheet = ws.max_column
        max_excel_col_index = column_index_from_string('XFD')

        for idx, cell_props in enumerate(footer_cell_data):
            target_col_idx = start_column_index + idx # Calculate target column index

            if target_col_idx > max_excel_col_index:
                 logging.warning(f"Calculated target column {target_col_idx} exceeds Excel limits ({max_excel_col_index}). Stopping row write.")
                 break

            try:
                target_cell = ws.cell(row=target_row_index, column=target_col_idx)

                # --- MODIFICATION ---
                # Apply border only if the target column is within the max border index determined from reference rows
                should_apply_border = (max_border_col_index > 0 and target_col_idx <= max_border_col_index)
                # --- END MODIFICATION ---

                apply_cell_properties(ws, target_cell, cell_props, should_apply_border)

            except Exception as cell_err:
                 logging.error(f"Failed to process cell properties for index {idx} (target col {target_col_idx}): {cell_err}")

        # Clear any remaining cells in the target row beyond the applied footer area
        # Clearing starts after the last cell written from footer_cell_data, regardless of bordering.
        clear_start_col = start_column_index + num_footer_cells

        if clear_start_col <= max_col_in_sheet:
             logging.debug(f"Clearing cells from column {clear_start_col} to {max_col_in_sheet} in row {target_row_index}")
             for col_clear_idx in range(clear_start_col, max_col_in_sheet + 1):
                 if col_clear_idx > max_excel_col_index:
                     logging.warning(f"Skipping clear for column {col_clear_idx} as it exceeds Excel limits.")
                     break
                 try:
                     cell_to_clear = ws.cell(row=target_row_index, column=col_clear_idx)
                     if cell_to_clear.value is not None or cell_to_clear.has_style:
                         cell_to_clear.value = None
                         if cell_to_clear.border != no_border:
                             cell_to_clear.border = no_border
                 except Exception as clear_err:
                      logging.error(f"Failed to clear cell at row {target_row_index}, col {col_clear_idx}: {clear_err}")

        logging.info(f"Successfully overwrote row {target_row_index} with footer data starting AT column {start_column_index}.")

    except Exception as e:
        logging.error(f"Failed to overwrite footer row at index {target_row_index}: {e}", exc_info=True)

# --- Main Execution Logic ---
def main(json_path: str, target_workbook_path: str, row_marker: str, col_pattern: str):
    """Loads data, finds markers, determines border extent, overwrites row."""
    logging.info(f"--- Starting Footer Overwrite using JSON: {json_path} ---")
    logging.info(f"--- Target Workbook: {target_workbook_path} ---")
    logging.info(f"--- Row Marker: '{row_marker}' ---")
    logging.info(f"--- Column Pattern: '{col_pattern}' ---")
    logging.info(f"--- Reference Row Markers for Bordering: {REFERENCE_ROW_MARKERS} ---")

    # --- 1. Load Extracted Footer Data ---
    if not os.path.exists(json_path):
        logging.critical(f"CRITICAL ERROR: Input JSON file not found at '{json_path}'.")
        return

    footer_cell_data = []
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            all_extracted_data = json.load(f)
        logging.info(f"Successfully loaded extracted data from '{json_path}'.")
        if not isinstance(all_extracted_data, dict) or not all_extracted_data:
             raise ValueError("Loaded JSON data is not a non-empty dictionary as expected.")
        first_sheet_name = next(iter(all_extracted_data))
        logging.info(f"Attempting to use footer data associated with key/sheet '{first_sheet_name}' from JSON.")
        footer_cell_data = all_extracted_data[first_sheet_name]
        if not isinstance(footer_cell_data, list):
             logging.warning(f"Data associated with key '{first_sheet_name}' in JSON is not a list. Footer might be empty.")
             footer_cell_data = []
        elif not footer_cell_data:
             logging.warning(f"Footer data list for key '{first_sheet_name}' in JSON is empty.")
        else:
             logging.info(f"Using footer data from key '{first_sheet_name}' found in JSON ({len(footer_cell_data)} cells).")

    except (json.JSONDecodeError, ValueError, StopIteration, KeyError) as load_err:
        logging.critical(f"CRITICAL ERROR: Failed to load or parse valid data from '{json_path}': {load_err}")
        return
    except Exception as load_err:
        logging.critical(f"CRITICAL ERROR: An unexpected error occurred loading data from '{json_path}': {load_err}", exc_info=True)
        return

    # --- 2. Load Target Workbook & Find Markers ---
    if not os.path.exists(target_workbook_path):
        logging.critical(f"CRITICAL ERROR: Target workbook not found at '{target_workbook_path}'.")
        return

    workbook = None
    target_row_idx = None
    start_col_idx = None
    target_sheet_name = None
    max_border_col_index = 0 # Initialize max column for bordering

    try:
        logging.info(f"Loading target workbook '{target_workbook_path}' for modification...")
        workbook = openpyxl.load_workbook(target_workbook_path)
        worksheet = workbook.active
        target_sheet_name = worksheet.title
        logging.info(f"Targeting active sheet: '{target_sheet_name}'")

        # --- 2a. Find Max Border Column from Reference Rows ---
        logging.info("Searching for reference rows to determine maximum border column...")
        reference_markers_lower = [m.lower() for m in REFERENCE_ROW_MARKERS]
        found_reference = False
        for r_idx in range(1, worksheet.max_row + 1):
            # Check cell in Column A (index 1)
            cell_A = worksheet.cell(row=r_idx, column=1)
            if cell_A.value is not None:
                try:
                    cell_value_lower = str(cell_A.value).strip().lower()
                    # Check if the cell value matches any of the reference markers
                    if cell_value_lower in reference_markers_lower:
                        logging.debug(f"Found reference row marker '{cell_A.value}' in row {r_idx}.")
                        found_reference = True
                        last_col_in_ref_row = find_last_col_with_value(worksheet, r_idx)
                        if last_col_in_ref_row > max_border_col_index:
                            logging.debug(f"Updating max border column from {max_border_col_index} to {last_col_in_ref_row} based on row {r_idx}.")
                            max_border_col_index = last_col_in_ref_row
                except Exception as check_err:
                     logging.warning(f"Could not check cell A{r_idx} value '{cell_A.value}': {check_err}")
                     continue # Skip to next row if check fails

        if not found_reference:
             logging.warning(f"Could not find any reference rows marked with {REFERENCE_ROW_MARKERS} in Column A. Borders may not be applied as expected.")
        elif max_border_col_index == 0:
             logging.warning(f"Found reference rows, but they appear to be empty. Borders will not be applied.")
        else:
             logging.info(f"Determined maximum column for bordering based on reference rows: {max_border_col_index}")
        # --- End Find Max Border Column ---


        # --- 2b. Find Target Row and Start Column for Footer ---
        target_row_idx = find_marker_row(worksheet, row_marker, MARKER_SEARCH_MAX_COL)
        if target_row_idx is None:
            logging.critical(f"CRITICAL ERROR: Row marker '{row_marker}' not found in sheet '{target_sheet_name}'. Cannot overwrite footer.")
            return

        start_col_idx = find_start_column(worksheet, target_row_idx, col_pattern, MARKER_SEARCH_MAX_COL)

        logging.info(f"Target row for footer found: {target_row_idx}. Starting column found/defaulted: {start_col_idx}. Proceeding with overwrite.")

        # --- 3. Overwrite Row with Footer Data ---
        overwrite_row_with_footer(worksheet, target_row_idx, start_col_idx, footer_cell_data, max_border_col_index)

        # --- 4. Save Modified Workbook ---
        logging.info(f"Attempting to save changes to '{target_workbook_path}'...")
        workbook.save(target_workbook_path)
        logging.info(f"--- Workbook saved successfully: '{target_workbook_path}' ---")
        print(f"\nFooter data processed for row {target_row_idx} starting at column {start_col_idx}. Workbook saved: {os.path.abspath(target_workbook_path)}")

    except PermissionError:
         logging.error(f"\n--- ERROR saving workbook: Permission denied. Is '{target_workbook_path}' open in another application? ---")
         print(f"\nERROR: Could not save the workbook. Please ensure '{os.path.basename(target_workbook_path)}' is closed and you have permission to write to it.")
    except FileNotFoundError:
         logging.critical(f"CRITICAL ERROR: Target workbook path suddenly invalid: '{target_workbook_path}'.")
         print(f"\nERROR: The path to the workbook '{target_workbook_path}' seems to be invalid.")
    except Exception as e:
        logging.critical(f"An unexpected error occurred during workbook processing or saving: {e}", exc_info=True)
        print(f"\nAn unexpected error occurred: {e}")
    finally:
        if workbook:
            try:
                workbook.close()
                logging.info("Target workbook closed.")
            except Exception as close_err:
                 logging.warning(f"Error closing target workbook object: {close_err}")

    logging.info("--- Script Finished ---")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Overwrite a marker row in an Excel file with footer data from JSON. Borders are applied based on reference rows.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
        )
    parser.add_argument("target_workbook",
                        help="Path to the target Excel workbook file to modify.")
    parser.add_argument("-j", "--json-input",
                        default=DEFAULT_INPUT_JSON,
                        help="Path to the input JSON file containing footer properties.")
    parser.add_argument("-r", "--row-marker",
                        default=DEFAULT_ROW_MARKER_TEXT,
                        help="The exact text marker identifying the row to overwrite.")
    parser.add_argument("-c", "--col-pattern",
                        default=DEFAULT_COL_MARKER_PATTERN,
                        help="The regex pattern to find the starting column within the target row (case-insensitive).")

    args = parser.parse_args()

    if not os.path.isfile(args.target_workbook):
         print(f"Error: Target workbook not found at '{args.target_workbook}'")
         logging.critical(f"CRITICAL ERROR: Target workbook not found at '{args.target_workbook}'")
    elif not os.path.isfile(args.json_input):
         print(f"Error: JSON input file not found at '{args.json_input}'")
         logging.critical(f"CRITICAL ERROR: JSON input file not found at '{args.json_input}'")
    else:
        main(args.json_input, args.target_workbook, args.row_marker, args.col_pattern)
