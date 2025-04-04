# total_row_value_extracter.py
# Finds a 'footer' row containing specific keywords (e.g., "Total", "Amount")
# and extracts cell values OR formulas from that row, mapped by sheet name.

import logging
import os
import json
import pprint
from typing import Dict, List, Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --- Configuration Constants ---
# !!! PLEASE MODIFY THESE VALUES !!!
WORKBOOK_PATH = r"test_inv - Copy.xlsx" # <<< CHANGE THIS to the actual path of your Excel file
# <<< Keywords to identify the footer row (case-insensitive contains check)
FOOTER_KEYWORDS = ["Total", "Amount"]
# --- End Configuration Constants ---

# --- Script Constants ---
# Optional: Start searching for the footer keywords from this row downwards
SEARCH_START_ROW = 5
# Optional: Limit the search for keywords to the first N columns for efficiency
SEARCH_MAX_COL = 6

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def find_footer_row(worksheet: Worksheet, keywords: List[str], start_row: int, max_col_to_search: int) -> Optional[int]:
    """
    Finds the 1-based index of the first row (at or after start_row)
    where any cell within the first few columns contains any of the specified keywords.
    NOTE: This searches the raw cell value, which may be a formula string if data_only=False.

    Args:
        worksheet: The openpyxl Worksheet object (loaded with data_only=False recommended).
        keywords: A list of strings to search for (case-insensitive contains check).
        start_row: The row number to begin searching from.
        max_col_to_search: The maximum column number to check in each row.

    Returns:
        The 1-based row index if a keyword is found, otherwise None.
    """
    keywords_lower = [k.lower() for k in keywords] # Prepare lowercase keywords for comparison
    logging.debug(f"Searching for footer row containing any of {keywords} in sheet '{worksheet.title}' (from row {start_row}, checking raw values/formulas)...")
    max_col = min(max_col_to_search, worksheet.max_column)

    # Iterate downwards from the specified start row
    for row_idx in range(start_row, worksheet.max_row + 1):
        # Check only the first few columns for the keywords
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # cell.value will contain formula if data_only=False
            if cell.value is not None:
                try:
                    # Convert raw value (could be formula) to string for searching
                    cell_value_str_lower = str(cell.value).lower().strip()
                    # Check if any keyword is present in the cell value/formula string
                    for keyword in keywords_lower:
                        if keyword in cell_value_str_lower:
                            logging.debug(f"Found keyword '{keyword}' (from {keywords}) in cell {cell.coordinate} (Raw Value: '{cell.value}', Row: {row_idx}). Identifying as footer row.")
                            return row_idx # Return the 1-based row index
                except Exception as e:
                    logging.warning(f"Could not check cell {cell.coordinate} value '{cell.value}': {e}")
                    continue # Skip cell on error
    # If loop completes without finding any keyword
    logging.warning(f"No keywords {keywords} found in any cell (raw value/formula) in sheet '{worksheet.title}' from row {start_row} onwards within first {max_col} columns.")
    return None

def extract_row_values(worksheet: Worksheet, target_row_index: int) -> List[Any]:
    """
    Extracts values OR formulas for all cells in the specified target row.
    Requires worksheet to be loaded with data_only=False to get formulas.

    Args:
        worksheet: The openpyxl Worksheet object (loaded with data_only=False).
        target_row_index: The 1-based index of the target row.

    Returns:
        A list containing the values or formula strings of the cells in the target row.
        Returns an empty list if target_row_index is invalid or row is empty.
    """
    row_cell_values: List[Any] = []
    if target_row_index is None or target_row_index < 1 or target_row_index > worksheet.max_row:
        logging.error(f"Invalid target_row_index ({target_row_index}) provided for value/formula extraction in sheet '{worksheet.title}'.")
        return row_cell_values

    logging.debug(f"Extracting values/formulas for all cells in row {target_row_index} from sheet '{worksheet.title}'.")

    # Iterate through ALL columns in the target row
    for col_idx in range(1, worksheet.max_column + 1):
        try:
            cell = worksheet.cell(row=target_row_index, column=col_idx)
            # Directly append the cell's value. If data_only=False, this will be
            # the formula string if one exists, otherwise the static value.
            row_cell_values.append(cell.value)

        except Exception as e:
            # Log the error and append a placeholder (e.g., None or an error string)
            logging.error(f"Error extracting value/formula for cell at row {target_row_index}, col {col_idx} in sheet '{worksheet.title}': {e}", exc_info=False)
            row_cell_values.append(None) # Append None as a placeholder for the failed cell
            continue # Continue to the next column

    logging.debug(f"Extracted {len(row_cell_values)} cell values/formulas from row {target_row_index}.")
    return row_cell_values

# --- Main Execution ---
if __name__ == "__main__":
    logging.info(f"--- Starting Footer Row Value/Formula Extraction (Keywords: {FOOTER_KEYWORDS}) in: {WORKBOOK_PATH} ---")

    # Check if the workbook exists
    if not os.path.exists(WORKBOOK_PATH):
        logging.critical(f"CRITICAL ERROR: Workbook not found at '{WORKBOOK_PATH}'. Please check the path.")
        exit(1)

    # Dictionary to hold results: { sheet_name: List[Any] }
    all_extracted_footer_row_values: Dict[str, List[Any]] = {}
    workbook = None

    try:
        # --- Load workbook with data_only=False to get formulas instead of calculated values ---
        logging.info("Loading workbook (data_only=False to capture formulas)...")
        # Set data_only=False here!
        workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
        logging.info(f"Workbook loaded. Sheets found: {workbook.sheetnames}")

        # Loop through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            logging.info(f"Processing sheet: '{sheet_name}'...")
            worksheet = workbook[sheet_name]

            # Find the footer row in the current sheet using the keywords
            # Note: find_footer_row now searches the raw cell value (which could be a formula)
            footer_row_index = find_footer_row(
                worksheet,
                FOOTER_KEYWORDS,
                SEARCH_START_ROW,
                SEARCH_MAX_COL
            )

            if footer_row_index is not None:
                logging.info(f"Footer row identified containing one of {FOOTER_KEYWORDS} at index {footer_row_index} in sheet '{sheet_name}'. Extracting row values/formulas...")
                # Extract values/formulas for all cells in that footer row
                row_values = extract_row_values(worksheet, footer_row_index) # This function now gets formulas due to load setting
                if row_values:
                    all_extracted_footer_row_values[sheet_name] = row_values
                    logging.info(f"Successfully extracted {len(row_values)} cell values/formulas from footer row in sheet '{sheet_name}'.")
                else:
                     logging.warning(f"Extraction returned no values/formulas for sheet '{sheet_name}' despite finding footer row index.")
            else:
                # Log if no footer keywords were found in this sheet
                logging.info(f"No row found containing keywords {FOOTER_KEYWORDS} (in raw cell value/formula) in sheet '{sheet_name}'. Skipping extraction for this sheet.")

    except Exception as e:
        logging.critical(f"An unexpected error occurred during processing: {e}", exc_info=True)
    finally:
        # --- Ensure workbook is closed properly ---
        if workbook:
            try:
                workbook.close()
                logging.info("Workbook closed.")
            except Exception as close_err:
                 logging.warning(f"Error closing workbook object: {close_err}")

    # --- Output Results ---
    logging.info("--- Extraction Complete ---")
    if all_extracted_footer_row_values:
        logging.info(f"Extracted footer row values/formulas from {len(all_extracted_footer_row_values)} sheet(s).")
        # Pretty print the results
        print(f"\n--- Extracted Values/Formulas for Footer Row (Keywords: {FOOTER_KEYWORDS}) ---")
        pprint.pprint(all_extracted_footer_row_values, indent=2, width=120)

        # Optionally save to a file (e.g., JSON)
        output_filename = f"extracted_footer_row_values_with_formulas.json" # Updated filename
        try:
            with open(output_filename, 'w', encoding='utf-8') as f:
                json.dump(all_extracted_footer_row_values, f, ensure_ascii=False, indent=4)
            logging.info(f"Results saved to {output_filename}")
        except TypeError as json_type_err:
             logging.error(f"Failed to save results to JSON due to non-serializable data type: {json_type_err}. Consider adding a default handler to json.dump().")
        except Exception as json_e:
            logging.error(f"Failed to save results to JSON: {json_e}")
    else:
        logging.warning(f"No footer rows (containing {FOOTER_KEYWORDS} in raw value/formula) were found and processed in any sheet.")

    logging.info("--- Script Finished ---")